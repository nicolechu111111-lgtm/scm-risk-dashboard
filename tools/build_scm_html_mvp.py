from __future__ import annotations

import json
import math
import os
from collections import defaultdict
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path

from openpyxl import load_workbook


TODAY = datetime.strptime(os.environ.get("SCM_TODAY", "2026-06-22"), "%Y-%m-%d").date()
WORKBOOK = Path(os.environ.get("SCM_WORKBOOK", "/Users/blue/Documents/跨境供应链跟单/outputs/scm_formula_template_20260620_v12/SCM Follow Up - Formula Template v12 Allocated ATP 2026-06-20.xlsx"))
OUT_DIR = Path(os.environ.get("SCM_OUT_DIR", "/Users/blue/Documents/跨境供应链跟单/outputs/scm_html_mvp_20260620"))
HTML_OUT = OUT_DIR / "SCM Risk Dashboard MVP 2026-06-20.html"
JSON_OUT = OUT_DIR / "scm_risk_dashboard_data_2026-06-20.json"


def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text or text in {"/", "-", "N/A", "NA"}:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


def as_num(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return 0
        return float(value)
    try:
        text = str(value).strip().replace(",", "")
        if not text or text in {"/", "-", "N/A", "NA"}:
            return 0
        return float(text)
    except Exception:
        return 0


def clean(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"/", "-", "None", "none"} else text


def header_key(value):
    return " ".join(clean(value).split()).casefold()


def fdate(value):
    return value.isoformat() if isinstance(value, date) else ""


def customer_name(value):
    text = clean(value)
    if "petco" in text.lower():
        return "Petco"
    if "petsmart" in text.lower():
        return "Petsmart"
    return text or "Unknown"


def is_excluded_sales_order(*values):
    text = " ".join(clean(value).lower() for value in values if clean(value))
    if "canada" in text:
        return True
    if "petco" in text and ("mx" in text or "mexico" in text):
        return True
    return any(token in text for token in {"petco-mx", "petco mx", "mx order", "mexico order"})


def transit_days(customer, dc):
    if "petsmart" in customer.lower():
        return 0
    text = (dc or "").upper()
    dc_defaults = {
        "MIRA LOMA": 2,
        "RENO": 3,
        "TEXAS": 5,
        "JOLIET": 7,
        "BRASELTON": 8,
        "CRANBURY": 9,
        "PSP DISTRIBUTION-SE": 7,
        "PSP DISTRIBUTION-OR": 8,
    }
    for key, days in dc_defaults.items():
        if key in text:
            return days
    states = {
        "CA": 2,
        "NV": 3,
        "AZ": 3,
        "OR": 4,
        "WA": 4,
        "UT": 5,
        "CO": 5,
        "TX": 5,
        "IL": 7,
        "GA": 8,
        "SC": 8,
        "NC": 8,
        "NJ": 9,
        "PA": 9,
        "NY": 9,
        "MD": 9,
        "VA": 9,
        "MA": 10,
        "CT": 10,
        "FL": 10,
    }
    for state, days in states.items():
        if f", {state}" in text or text.endswith(f" {state}") or text.endswith(state):
            return days
    return 7


def status_rank(status):
    return {"Critical Risk": 0, "Warehouse Prep Risk": 1, "Planned PO Watch": 2, "Lead Time Watch": 3, "OK": 4}.get(status, 5)


def rowdict(headers, row):
    return {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


def header_index(ws):
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {header_key(value): i for i, value in enumerate(headers) if clean(value)}


def col(headers, *names, default=None):
    for name in names:
        idx = headers.get(header_key(name))
        if idx is not None:
            return idx
    return default


def value_at(row, idx, default=None):
    if idx is None or idx >= len(row):
        return default
    return row[idx]


def main():
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws_sum = wb["Sum"]
    ws_sales = wb["Sales"]
    ws_purchase = wb["Purchase"]
    ws_shipment = wb["Shipment Control"] if "Shipment Control" in wb.sheetnames else None
    ws_product = wb["Product List"]
    ws_customer_code = wb["Customer Code"]

    sum_cols = header_index(ws_sum)
    sales_cols = header_index(ws_sales)
    purchase_cols = header_index(ws_purchase)
    shipment_cols = header_index(ws_shipment) if ws_shipment else {}
    product_cols = header_index(ws_product)
    customer_code_cols = header_index(ws_customer_code)

    base_date = as_date(ws_sum["C1"].value) or date(2026, 4, 13)

    product_name = {}
    product_cost = {}
    lead_weeks = {}
    product_upc = {}
    case_pack_by_sku = {}
    upc_to_sku = {}
    product_sku_col = col(product_cols, "Item Code", default=0)
    product_upc_col = col(product_cols, "UPC", default=1)
    product_name_col = col(product_cols, "Product Name", default=2)
    product_cost_col = col(product_cols, "Cost", default=8)
    product_lt_col = col(product_cols, "LT", default=6)
    product_case_pack_col = col(product_cols, "Case Pack", default=7)
    for row in ws_product.iter_rows(min_row=2, values_only=True):
        sku = clean(value_at(row, product_sku_col))
        if not sku:
            continue
        product_name[sku] = clean(value_at(row, product_name_col))
        product_cost[sku] = as_num(value_at(row, product_cost_col))
        lead_weeks[sku] = as_num(value_at(row, product_lt_col)) or 10
        product_upc[sku] = clean(value_at(row, product_upc_col)).split(".")[0]
        case_pack_by_sku[sku] = int(as_num(value_at(row, product_case_pack_col)))
        upc = clean(value_at(row, product_upc_col))
        if upc:
            upc_to_sku[upc.split(".")[0]] = sku

    customer_code_to_sku = {}
    customer_code_to_upc = {}
    customer_code_col = col(customer_code_cols, "Customer Code", default=1)
    customer_sku_col = col(customer_code_cols, "Item Code", default=2)
    customer_upc_col = col(customer_code_cols, "UPC", default=3)
    for row in ws_customer_code.iter_rows(min_row=2, values_only=True):
        code = clean(value_at(row, customer_code_col))
        sku = clean(value_at(row, customer_sku_col))
        if code and sku:
            customer_code_to_sku[code.split(".")[0]] = sku
            customer_code_to_upc[code.split(".")[0]] = clean(value_at(row, customer_upc_col)).split(".")[0]

    base_stock = {}
    weekly_forecast_by_sku = {}
    sum_lt_by_sku = {}
    safety_stock_by_sku = {}
    moq_by_sku = {}
    total_stock_by_sku = {}
    sum_case_pack_by_sku = {}
    purchase_suggestion_by_sku = {}
    sum_sku_col = col(sum_cols, "Item Code", default=0)
    sum_onhand_col = col(sum_cols, "Onhand Stock", default=7 if ws_sum.max_column > 7 else 2)
    sum_total_stock_col = col(sum_cols, "Total Stock", default=None)
    sum_forecast_col = col(sum_cols, "Weekly Forecast", default=None)
    sum_lt_col = col(sum_cols, "LT", default=None)
    sum_ss_col = col(sum_cols, "SS", default=None)
    sum_moq_col = col(sum_cols, "MOQ", default=None)
    sum_purchase_suggestion_col = col(sum_cols, "Purchase Suggestion", default=None)
    sum_case_pack_col = col(sum_cols, "Case Pack", default=None)
    for row in ws_sum.iter_rows(min_row=2, values_only=True):
        sku = clean(value_at(row, sum_sku_col))
        if sku:
            base_stock[sku] = as_num(value_at(row, sum_onhand_col))
            total_stock_by_sku[sku] = as_num(value_at(row, sum_total_stock_col))
            weekly_forecast_by_sku[sku] = as_num(value_at(row, sum_forecast_col))
            sum_lt_by_sku[sku] = as_num(value_at(row, sum_lt_col)) or lead_weeks.get(sku, 10)
            safety_stock_by_sku[sku] = as_num(value_at(row, sum_ss_col))
            moq_by_sku[sku] = as_num(value_at(row, sum_moq_col))
            purchase_suggestion_by_sku[sku] = as_num(value_at(row, sum_purchase_suggestion_col))
            sum_case_pack_by_sku[sku] = as_num(value_at(row, sum_case_pack_col))

    future_pos_by_sku = defaultdict(list)
    purchase_rows = []
    arrived_po_rows = []
    shipment_control_rows = []
    shipment_by_bol = {}
    if ws_shipment:
        shipment_bol_col = col(shipment_cols, "BOL", default=0)
        shipment_factory_col = col(shipment_cols, "Factory Delivery Date", default=4)
        shipment_sailing_col = col(shipment_cols, "Sailing Date", default=5)
        shipment_port_eta_col = col(shipment_cols, "ETD（到港）", "ETD(到港)", "ETD", default=6)
        shipment_port_atd_col = col(shipment_cols, "ATD（到港）", "ATD(到港)", "ATD", default=7)
        for i, row in enumerate(ws_shipment.iter_rows(min_row=2, values_only=True), start=2):
            bol = clean(value_at(row, shipment_bol_col))
            if not bol:
                continue
            item = {
                "row": i,
                "bol": bol,
                "factory_date": fdate(as_date(value_at(row, shipment_factory_col))),
                "sailing_date": fdate(as_date(value_at(row, shipment_sailing_col))),
                "port_eta": fdate(as_date(value_at(row, shipment_port_eta_col))),
                "port_atd": fdate(as_date(value_at(row, shipment_port_atd_col))),
            }
            shipment_control_rows.append(item)
            shipment_by_bol[bol] = item
    purchase_sku_col = col(purchase_cols, "Product Code", default=2)
    purchase_po_col = col(purchase_cols, "PO/PI", default=4)
    purchase_qty_col = col(purchase_cols, "ORD QTY", default=6)
    purchase_factory_col = col(purchase_cols, "Factory Delivery Date", default=7)
    purchase_sailing_col = col(purchase_cols, "Sailing Date", default=8)
    purchase_port_eta_col = col(purchase_cols, "ETD（到港）", "ETD(到港)", "ETD", default=9)
    purchase_receipt_col = col(purchase_cols, "ATD（到仓）", "ATD（到港）", "ATD", default=10)
    purchase_bol_col = col(purchase_cols, "BOL", default=16)
    for i, row in enumerate(ws_purchase.iter_rows(min_row=2, values_only=True), start=2):
        sku = clean(value_at(row, purchase_sku_col))
        if not sku:
            continue
        po = clean(value_at(row, purchase_po_col)) or "(blank PO)"
        qty = as_num(value_at(row, purchase_qty_col))
        factory_date = as_date(value_at(row, purchase_factory_col))
        sailing_date = as_date(value_at(row, purchase_sailing_col))
        port_eta = as_date(value_at(row, purchase_port_eta_col))
        actual_receipt = as_date(value_at(row, purchase_receipt_col))
        bol = clean(value_at(row, purchase_bol_col))
        shipment_row = shipment_by_bol.get(bol, {}) if bol else {}
        shipment_port_atd = as_date(shipment_row.get("port_atd"))

        available_date = None
        reliability = ""
        port_based_available = port_eta + timedelta(days=7) if port_eta else None
        is_received = bool(actual_receipt)
        if is_received:
            reliability = "已到货"
            arrived_po_rows.append(
                {
                    "row": i,
                    "sku": sku,
                    "product": product_name.get(sku, ""),
                    "po": po,
                    "qty": int(qty),
                    "actual_receipt_date": fdate(actual_receipt),
                    "sailing_date": fdate(sailing_date),
                    "port_eta": fdate(port_eta),
                    "bol": bol,
                    "status": "已到货，已从未来在途排除",
                }
            )
        elif port_based_available:
            available_date = port_based_available
            if sailing_date and port_eta and port_eta <= TODAY:
                reliability = "已到港待送仓"
            else:
                reliability = "确认在途" if sailing_date else "预计在途"
        elif factory_date and factory_date > TODAY:
            reliability = "未发货-需确认工厂交期"

        entry = {
            "row": i,
            "sku": sku,
            "product": product_name.get(sku, ""),
            "po": po,
            "qty": int(qty),
            "factory_date": fdate(factory_date),
            "sailing_date": fdate(sailing_date),
            "port_eta": fdate(port_eta),
            "available_date": fdate(available_date),
            "bol": bol,
            "actual_receipt_date": fdate(actual_receipt),
            "shipment_port_atd": fdate(shipment_port_atd),
            "reliability": reliability or "无可用到仓日期",
        }
        purchase_rows.append(entry)
        if not is_received and available_date and qty > 0:
            future_pos_by_sku[sku].append(
                {
                    "po": po,
                    "qty": qty,
                    "remaining": qty,
                    "available": available_date,
                    "reliability": reliability,
                    "sailing_date": sailing_date,
                    "port_eta": port_eta,
                    "factory_date": factory_date,
                    "bol": bol,
                    "actual_receipt_date": actual_receipt,
                    "shipment_port_atd": shipment_port_atd,
                }
            )

    sales_lines = []
    old_open_lines = []
    excluded_sales_lines = []
    shipped_new_by_sku = defaultdict(float)
    sales_order_date_col = col(sales_cols, "Date", default=0)
    sales_sku_col = col(sales_cols, "Product Code", default=2)
    sales_so_col = col(sales_cols, "SO/CI", default=3)
    sales_qty_col = col(sales_cols, "ORD QTY", default=5)
    sales_customer_col = col(sales_cols, "Customer", default=8)
    sales_dc_col = col(sales_cols, "Delivery Center", default=9)
    sales_required_col = col(sales_cols, "ETD", default=10)
    sales_delivered_col = col(sales_cols, "ATD", default=11)
    for i, row in enumerate(ws_sales.iter_rows(min_row=2, values_only=True), start=2):
        order_date = as_date(value_at(row, sales_order_date_col))
        sku = clean(value_at(row, sales_sku_col))
        so = clean(value_at(row, sales_so_col))
        qty = as_num(value_at(row, sales_qty_col))
        raw_customer = clean(value_at(row, sales_customer_col))
        customer = customer_name(raw_customer)
        dc = clean(value_at(row, sales_dc_col))
        required_arrival = as_date(value_at(row, sales_required_col))
        delivered_at = as_date(value_at(row, sales_delivered_col))
        if not sku or not so or qty <= 0:
            continue

        if is_excluded_sales_order(raw_customer, so, dc):
            if not delivered_at:
                excluded_sales_lines.append(
                    {
                        "row": i,
                        "so": so,
                        "sku": sku,
                        "qty": int(qty),
                        "customer": raw_customer,
                        "delivery_center": dc,
                        "required_arrival": fdate(required_arrival),
                        "reason": "Petco MX/Canada 不参与 US 发货风险和补货计算",
                    }
                )
            continue

        if order_date and order_date > base_date and delivered_at:
            shipped_new_by_sku[sku] += qty

        if delivered_at:
            continue

        if not order_date or order_date <= base_date:
            old_open_lines.append({"row": i, "so": so, "sku": sku, "qty": int(qty), "order_date": fdate(order_date), "required_arrival": fdate(required_arrival)})
            continue

        days = transit_days(customer, dc)
        buffer_days = 1
        latest_ship = required_arrival - timedelta(days=days + buffer_days) if required_arrival else None
        sales_lines.append(
            {
                "row": i,
                "order_date": order_date,
                "sku": sku,
                "product": product_name.get(sku, ""),
                "so": so,
                "qty": qty,
                "customer": customer,
                "delivery_center": dc,
                "required_arrival": required_arrival,
                "domestic_transit": days,
                "latest_ship": latest_ship,
                "lead_weeks": lead_weeks.get(sku, 10),
            }
        )

    by_sku = defaultdict(list)
    for line in sales_lines:
        by_sku[line["sku"]].append(line)

    def estimate_so_issue_profile():
        """先跑一遍粗分配，判断每张 SO 本来有几个问题 SKU。

        第二轮正式分配时会用这个结果优先保整单：同一紧急窗口里，
        问题 SKU 少的 SO 先拿库存/在途，问题 SKU 多的 SO 更适合作为集中沟通对象。
        """
        profile = defaultdict(lambda: {"issue_skus": set(), "critical_skus": set(), "issue_qty": 0})
        for sku, lines in by_sku.items():
            lots = [
                {"qty": float(base_stock.get(sku, 0)), "remaining": float(base_stock.get(sku, 0)), "available": date.min, "reliability": "现有库存"}
            ]
            lots.extend(
                {
                    "qty": float(po["qty"]),
                    "remaining": float(po["qty"]),
                    "available": po["available"],
                    "reliability": po["reliability"],
                }
                for po in future_pos_by_sku.get(sku, [])
            )
            lots = sorted(lots, key=lambda x: (x["available"], 1 if x["reliability"] in {"Planned ETA", "预计在途"} else 0))
            for line in sorted(lines, key=lambda x: (x["latest_ship"] or date.max, x["required_arrival"] or date.max, x["row"])):
                need = float(line["qty"])
                cover_date = None
                planned_cover = False
                for lot in lots:
                    if need <= 0:
                        break
                    if lot["remaining"] <= 0:
                        continue
                    take = min(lot["remaining"], need)
                    lot["remaining"] -= take
                    need -= take
                    cover_date = max(cover_date or lot["available"], lot["available"])
                    planned_cover = planned_cover or lot["reliability"] in {"Planned ETA", "预计在途"}
                latest_ship = line["latest_ship"]
                has_issue = need > 0 or bool(latest_ship and cover_date and cover_date > latest_ship) or planned_cover
                has_critical = need > 0 or bool(latest_ship and cover_date and cover_date > latest_ship)
                if has_issue:
                    item = profile[line["so"]]
                    item["issue_skus"].add(sku)
                    item["issue_qty"] += int(max(need, line["qty"]))
                    if has_critical:
                        item["critical_skus"].add(sku)
        return {
            so: {
                "issue_skus": len(item["issue_skus"]),
                "critical_skus": len(item["critical_skus"]),
                "issue_qty": item["issue_qty"],
            }
            for so, item in profile.items()
        }

    so_issue_profile = estimate_so_issue_profile()

    def allocation_sort_key(line):
        ship = line["latest_ship"] or date.max
        required = line["required_arrival"] or date.max
        profile = so_issue_profile.get(line["so"], {"issue_skus": 0, "critical_skus": 0, "issue_qty": 0})
        is_overdue = bool(line["latest_ship"] and line["latest_ship"] < TODAY)
        # 7 天为一个紧急窗口：窗口内优先保整单，窗口外仍尊重客户交期。
        ship_window = ship.toordinal() // 7
        return (
            0 if is_overdue else 1,
            ship_window,
            profile["critical_skus"],
            profile["issue_skus"],
            -profile["issue_qty"],
            ship,
            required,
            line["row"],
        )

    sku_summary = []
    so_lines = []
    po_impact = defaultdict(lambda: {"po": "", "sku_count": set(), "so_count": set(), "critical": 0, "watch": 0, "late_qty": 0, "lines": []})

    for sku, lines in by_sku.items():
        # Sum Onhand 是当前真实库存；不要再用期初日期后的到货/出货重算，
        # 否则已到货 PO 或已出货 SO 会被重复处理。
        current_onhand = base_stock.get(sku, 0)
        remaining_stock = current_onhand
        pos = sorted(
            [dict(po) for po in future_pos_by_sku.get(sku, [])],
            key=lambda x: (x["available"], 1 if x["reliability"] in {"Planned ETA", "预计在途"} else 0, x["po"]),
        )
        sorted_lines = sorted(lines, key=allocation_sort_key)
        issue_count = critical_count = watch_count = planned_count = 0
        issue_qty = critical_qty = watch_qty = planned_qty = uncovered_qty = 0
        for line in sorted_lines:
            qty = line["qty"]
            used_stock = min(max(remaining_stock, 0), qty)
            remaining_stock -= used_stock
            need = qty - used_stock
            covers = []
            uncovered = 0
            if need > 0:
                for po in pos:
                    if po["remaining"] <= 0:
                        continue
                    take = min(po["remaining"], need)
                    po["remaining"] -= take
                    need -= take
                    covers.append(
                        {
                            "po": po["po"],
                            "qty": int(take),
                            "available": po["available"],
                            "reliability": po["reliability"],
                            "bol": po["bol"],
                        }
                    )
                    if need <= 0:
                        break
                uncovered = need

            cover_date = max((c["available"] for c in covers), default=None)
            warehouse_ready = cover_date + timedelta(days=7) if cover_date else None
            planned_cover = any(c["reliability"] in {"Planned ETA", "预计在途"} for c in covers)
            confirmed_cover = bool(covers) and not planned_cover
            latest_ship = line["latest_ship"]
            order_date = line["order_date"]
            lead_due = order_date + timedelta(days=int(line["lead_weeks"] * 7)) if order_date else None
            suggested_customer_date = fdate(warehouse_ready + timedelta(days=line["domestic_transit"] + 1)) if warehouse_ready else ""
            if uncovered > 0:
                status = "Critical Risk"
                action = f"Short {int(uncovered)} units. Need new/earlier PO."
            elif not covers:
                status = "OK"
                action = "Covered by current stock."
            elif latest_ship and cover_date and cover_date > latest_ship:
                status = "Critical Risk"
                action = f"Need earlier ETA for {covers[-1]['po']} by {(cover_date - latest_ship).days} days."
                if planned_cover:
                    action += " Cover is planned ETA, confirm sailing date."
            elif latest_ship and warehouse_ready and warehouse_ready > latest_ship:
                status = "Warehouse Prep Risk"
                action = f"Warehouse needs 7 days after arrival; suggest customer date {suggested_customer_date}."
            elif lead_due and latest_ship and lead_due > latest_ship:
                status = "Lead Time Watch"
                action = "Supplier lead time is tight; confirm PO timing."
            elif planned_cover:
                status = "Planned PO Watch"
                action = "Covered only if planned ETA holds; confirm sailing date before promising."
            else:
                status = "OK"
                action = "Covered by confirmed incoming PO on time."

            if status != "OK":
                issue_count += 1
                issue_qty += max(int(qty - used_stock), int(uncovered))
            if status == "Critical Risk":
                critical_count += 1
                critical_qty += max(int(qty - used_stock), int(uncovered))
            if status == "Lead Time Watch":
                watch_count += 1
                watch_qty += max(int(qty - used_stock), int(uncovered))
            if status == "Planned PO Watch":
                planned_count += 1
                planned_qty += max(int(qty - used_stock), int(uncovered))
            uncovered_qty += int(uncovered)

            cover_text = "Current stock" if not covers and uncovered == 0 else "; ".join(
                f"{c['po']}: {c['qty']} @ {fdate(c['available'])} ({c['reliability']})" for c in covers
            )
            if uncovered:
                cover_text = (cover_text + "; " if cover_text else "") + f"Uncovered {int(uncovered)}"

            out = {
                "status": status,
                "so": line["so"],
                "customer": line["customer"],
                "delivery_center": line["delivery_center"],
                "sku": sku,
                "product": line["product"],
                "qty": int(qty),
                "required_arrival": fdate(line["required_arrival"]),
                "latest_ship": fdate(latest_ship),
                "domestic_transit": line["domestic_transit"],
                "stock_before": int(max(remaining_stock + used_stock, 0)),
                "stock_used": int(used_stock),
                "cover": cover_text,
                "cover_pos": [c["po"] for c in covers],
                "cover_eta": fdate(cover_date),
                "warehouse_ready_date": fdate(warehouse_ready),
                "cover_type": "预计在途" if planned_cover else ("确认在途" if confirmed_cover else "现有库存"),
                "uncovered": int(uncovered),
                "risk_qty": max(int(qty - used_stock), int(uncovered)) if status != "OK" else 0,
                "action": action,
                "suggested_customer_date": suggested_customer_date if status == "Warehouse Prep Risk" else "",
                "row": line["row"],
                "lead_weeks": line["lead_weeks"],
                "lead_due": fdate(lead_due),
                "allocation_issue_skus": so_issue_profile.get(line["so"], {}).get("issue_skus", 0),
                "allocation_critical_skus": so_issue_profile.get(line["so"], {}).get("critical_skus", 0),
            }
            so_lines.append(out)
            for c in covers:
                item = po_impact[c["po"]]
                item["po"] = c["po"]
                item["sku_count"].add(sku)
                item["so_count"].add(line["so"])
                item["late_qty"] += int(c["qty"]) if status == "Critical Risk" else 0
                item["critical"] += 1 if status == "Critical Risk" else 0
                item["watch"] += 1 if status == "Lead Time Watch" or (status == "Planned PO Watch" and c["reliability"] not in {"Planned ETA", "预计在途"}) else 0
                item["planned"] = item.get("planned", 0) + (1 if status == "Planned PO Watch" and c["reliability"] in {"Planned ETA", "预计在途"} else 0)
                item["lines"].append({
                    "so": line["so"],
                    "sku": sku,
                    "product": product_name.get(sku, ""),
                    "qty": int(c["qty"]),
                    "available": fdate(c["available"]),
                    "latest_ship": fdate(latest_ship),
                    "days_delta": (c["available"] - latest_ship).days if latest_ship else "",
                    "status": status,
                    "reliability": c["reliability"],
                    "delivery_center": line["delivery_center"],
                })

        open_demand = int(sum(x["qty"] for x in sorted_lines))
        stock_gap = max(open_demand - max(int(current_onhand), 0), 0)
        future_qty = int(sum(x["qty"] for x in pos))
        sku_summary.append(
            {
                "sku": sku,
                "product": product_name.get(sku, ""),
                "current_onhand": int(current_onhand),
                "weekly_forecast": int(weekly_forecast_by_sku.get(sku, 0)),
                "lead_weeks": sum_lt_by_sku.get(sku, lead_weeks.get(sku, 10)),
                "safety_stock_weeks": safety_stock_by_sku.get(sku, 0),
                "moq": int(moq_by_sku.get(sku, 0)),
                "open_demand": open_demand,
                "stock_gap": stock_gap,
                "future_qty": future_qty,
                "net_gap_after_future": max(stock_gap - future_qty, 0),
                "issue_qty": int(issue_qty),
                "critical_qty": int(critical_qty),
                "watch_qty": int(watch_qty),
                "planned_qty": int(planned_qty),
                "uncovered_qty": int(uncovered_qty),
                "issue_count": issue_count,
                "critical_count": critical_count,
                "watch_count": watch_count,
                "planned_count": planned_count,
            }
        )

    so_all = defaultdict(lambda: {"so": "", "customer": "", "delivery_center": "", "required_arrival": "", "latest_ship": "", "line_count": 0, "issue_skus": 0, "critical_skus": 0, "warehouse_skus": 0, "watch_skus": 0, "planned_skus": 0, "confirmed_po_skus": 0, "total_qty": 0, "skus": []})
    for line in so_lines:
        item = so_all[line["so"]]
        item["so"] = line["so"]
        item["customer"] = line["customer"]
        item["delivery_center"] = line["delivery_center"]
        item["required_arrival"] = min([x for x in [item["required_arrival"], line["required_arrival"]] if x] or [line["required_arrival"]])
        item["latest_ship"] = min([x for x in [item["latest_ship"], line["latest_ship"]] if x] or [line["latest_ship"]])
        item["line_count"] += 1
        item["total_qty"] += line["qty"]
        item["skus"].append(line)
    for item in so_all.values():
        issue_lines = [x for x in item["skus"] if x["status"] != "OK"]
        item["issue_skus"] = len({x["sku"] for x in issue_lines})
        item["critical_skus"] = len({x["sku"] for x in issue_lines if x["status"] == "Critical Risk"})
        item["warehouse_skus"] = len({x["sku"] for x in issue_lines if x["status"] == "Warehouse Prep Risk"})
        item["watch_skus"] = len({x["sku"] for x in issue_lines if x["status"] == "Lead Time Watch"})
        item["planned_skus"] = len({x["sku"] for x in issue_lines if x["status"] == "Planned PO Watch"})
        item["confirmed_po_skus"] = len({x["sku"] for x in issue_lines if x.get("cover_type") == "Confirmed Incoming"})
        item["suggested_customer_date"] = max([x.get("suggested_customer_date", "") for x in issue_lines] or [""])
        item["status"] = "Critical Risk" if item["critical_skus"] else ("Warehouse Prep Risk" if item["warehouse_skus"] else ("Planned PO Watch" if item["planned_skus"] else ("Lead Time Watch" if item["watch_skus"] else "OK")))
        item["action"] = (
            "Ask sales/customer to revise date or expedite cover PO."
            if item["critical_skus"]
            else ("Warehouse needs 7 days after arrival; suggest customer date {}.".format(item["suggested_customer_date"]) if item["warehouse_skus"] else ("Confirm sailing date/factory timing for planned ETA PO." if item["planned_skus"] else ("Confirmed incoming PO covers; monitor ETA/receiving." if item["watch_skus"] else "No action.")))
        )

    so_board = sorted(so_all.values(), key=lambda x: (status_rank(x["status"]), x["latest_ship"] or "9999-12-31", -x["issue_skus"], x["so"]))
    critical_sos = [x for x in so_board if x["status"] == "Critical Risk"]
    watch_sos = [x for x in so_board if x["status"] in {"Warehouse Prep Risk", "Lead Time Watch", "Planned PO Watch"}]

    recommendations = []
    by_sku_issue = defaultdict(list)
    for line in so_lines:
        if line["status"] != "OK":
            by_sku_issue[line["sku"]].append(line)
    for sku, issue_lines in by_sku_issue.items():
        sku_orders = [x for x in so_lines if x["sku"] == sku]
        affected_so = sorted(issue_lines, key=lambda x: (status_rank(x["status"]), x["latest_ship"] or "9999-12-31"))[0]["so"]
        candidates = sorted(
            (so_all[x["so"]] for x in sku_orders),
            key=lambda so: (
                0 if so["critical_skus"] else 1,
                -so["issue_skus"],
                so["latest_ship"] or "9999-12-31",
                so["so"],
            ),
        )
        rec = candidates[0]
        recommendations.append(
            {
                "sku": sku,
                "product": product_name.get(sku, ""),
                "stock_gap": next((s["stock_gap"] for s in sku_summary if s["sku"] == sku), 0),
                "issue_qty": next((s["issue_qty"] for s in sku_summary if s["sku"] == sku), 0),
                "standard_affected_so": affected_so,
                "recommended_so": rec["so"],
                "recommended_reason": f"{rec['so']} has {rec['critical_skus']} critical SKU(s) and {rec['issue_skus']} issue SKU(s), so communication can be consolidated.",
                "recommendation": "Consider assigning shortage/date-change discussion here." if rec["so"] != affected_so else "Natural allocation already points here.",
            }
        )

    po_rows = []
    for item in po_impact.values():
        po_rows.append(
            {
                "po": item["po"],
                "sku_count": len(item["sku_count"]),
                "so_count": len(item["so_count"]),
                "critical_lines": item["critical"],
                "watch_lines": item["watch"],
                "planned_lines": item.get("planned", 0),
                "late_qty": item["late_qty"],
                "priority_lines": sorted(item["lines"], key=lambda x: (status_rank(x["status"]), x["latest_ship"], x["sku"], x["so"]))[:80],
            }
        )
    po_rows.sort(key=lambda x: (-x["critical_lines"], -x["planned_lines"], -x["watch_lines"], x["po"]))

    replenishment_rows = []
    all_skus = sorted(set(base_stock) | set(by_sku) | set(future_pos_by_sku))
    for sku in all_skus:
        onhand = int(base_stock.get(sku, 0))
        open_demand = int(sum(line["qty"] for line in by_sku.get(sku, [])))
        pos = future_pos_by_sku.get(sku, [])
        confirmed_incoming = int(sum(po["qty"] for po in pos if po["reliability"] not in {"Planned ETA", "预计在途"}))
        planned_incoming = int(sum(po["qty"] for po in pos if po["reliability"] in {"Planned ETA", "预计在途"}))
        weekly_forecast = int(weekly_forecast_by_sku.get(sku, 0))
        lead_time_weeks = int(sum_lt_by_sku.get(sku, lead_weeks.get(sku, 10)) or 0)
        safety_weeks = int(safety_stock_by_sku.get(sku, 0) or 0)
        moq = int(moq_by_sku.get(sku, 0) or 0)
        case_pack = int(sum_case_pack_by_sku.get(sku, 0) or case_pack_by_sku.get(sku, 0) or 0)
        total_stock = int(total_stock_by_sku.get(sku, 0))
        target_weeks = lead_time_weeks + safety_weeks + 2
        target_stock = int(weekly_forecast * max(target_weeks, 0))
        raw_gap = max(target_stock - total_stock, 0)
        rounded_qty = int(math.ceil(raw_gap / case_pack) * case_pack) if raw_gap and case_pack else int(raw_gap)
        sum_purchase_suggestion = int(purchase_suggestion_by_sku.get(sku, 0))
        order_risk_gap = max(open_demand - onhand - confirmed_incoming - planned_incoming, 0)
        status = "建议补货" if rounded_qty else ("观察" if weekly_forecast and total_stock - target_stock < weekly_forecast * 2 else "无需补货")
        replenishment_rows.append(
            {
                "status": status,
                "sku": sku,
                "product": product_name.get(sku, ""),
                "onhand": onhand,
                "open_demand": open_demand,
                "confirmed_incoming": confirmed_incoming,
                "planned_incoming": planned_incoming,
                "total_stock": total_stock,
                "weekly_forecast": weekly_forecast,
                "lead_time_weeks": lead_time_weeks,
                "safety_stock_weeks": safety_weeks,
                "buffer_weeks": 2,
                "target_weeks": target_weeks,
                "target_stock": target_stock,
                "formula_gap": raw_gap,
                "case_pack": case_pack,
                "moq": moq,
                "suggested_po_qty": rounded_qty,
                "sum_purchase_suggestion": sum_purchase_suggestion,
                "order_risk_gap": int(order_risk_gap),
                "action": (
                    f"建议下单 {rounded_qty}；公式缺口 {raw_gap}，已按箱规 {case_pack or '-'} 向上取整。"
                    if rounded_qty
                    else "Total Stock 可覆盖 LT+SS+2 周预测目标。"
                ),
            }
        )

    reminders = []
    for so in so_board:
        ship = as_date(so["latest_ship"])
        if ship and 0 <= (ship - TODAY).days <= 21 and so["status"] != "OK":
            reminders.append({k: so[k] for k in ["so", "customer", "delivery_center", "latest_ship", "required_arrival", "status", "issue_skus", "critical_skus", "action"]})

    action_queue = []
    seen_so = set()
    for so in so_board:
        ship = as_date(so["latest_ship"])
        days = (ship - TODAY).days if ship else None
        include = so["status"] == "Critical Risk" or (so["status"] in {"Warehouse Prep Risk", "Lead Time Watch", "Planned PO Watch"} and days is not None and days <= 21)
        if not include or so["so"] in seen_so:
            continue
        seen_so.add(so["so"])
        if so["status"] == "Critical Risk":
            reason = "Critical risk"
            if days is not None and days < 0:
                reason = "Overdue critical risk"
        elif so["status"] == "Warehouse Prep Risk":
            reason = "仓库备货风险"
        elif so["status"] == "Planned PO Watch":
            reason = "Planned ETA needs confirmation"
        else:
            reason = "21-day warehouse notice"
        action_queue.append({**{k: so[k] for k in ["so", "customer", "delivery_center", "latest_ship", "required_arrival", "status", "line_count", "issue_skus", "critical_skus", "watch_skus", "planned_skus", "confirmed_po_skus", "action"]}, "reason": reason, "days_to_ship": days if days is not None else ""})

    transit_check_rows = []
    seen_dc = set()
    for line in sales_lines:
        key = (line["customer"], line["delivery_center"])
        if key in seen_dc:
            continue
        seen_dc.add(key)
        dc_text = (line["delivery_center"] or "").upper()
        known = any(token in dc_text for token in ["MIRA LOMA", "RENO", "TEXAS", "JOLIET", "BRASELTON", "CRANBURY", "PSP DISTRIBUTION-SE", "PSP DISTRIBUTION-OR"])
        state_known = any(f", {st}" in dc_text or dc_text.endswith(f" {st}") or dc_text.endswith(st) for st in ["CA","NV","AZ","OR","WA","UT","CO","TX","IL","GA","SC","NC","NJ","PA","NY","MD","VA","MA","CT","FL"])
        if not known and not state_known and line["customer"] != "Petsmart":
            transit_check_rows.append(
                {
                    "customer": line["customer"],
                    "delivery_center": line["delivery_center"],
                    "current_days": transit_days(line["customer"], line["delivery_center"]),
                    "issue": "未识别客户仓州/默认按7天",
                    "action": "请到运输设置页维护该客户仓天数。",
                }
            )

    data = {
        "generated_at": TODAY.isoformat(),
        "base_date": base_date.isoformat(),
        "source_workbook": str(WORKBOOK),
        "customer_code_to_sku": customer_code_to_sku,
        "customer_code_to_upc": customer_code_to_upc,
        "upc_to_sku": upc_to_sku,
        "product_name_by_sku": product_name,
        "product_cost_by_sku": {sku: cost for sku, cost in product_cost.items() if cost},
        "product_upc_by_sku": {sku: upc for sku, upc in product_upc.items() if upc},
        "case_pack_by_sku": {sku: pack for sku, pack in case_pack_by_sku.items() if pack},
        "inventory_by_sku": {
            sku: {
                "sku": sku,
                "product": product_name.get(sku, ""),
                "current_onhand": int(qty),
                "weekly_forecast": int(weekly_forecast_by_sku.get(sku, 0)),
                "lead_weeks": sum_lt_by_sku.get(sku, lead_weeks.get(sku, 10)),
                "safety_stock_weeks": safety_stock_by_sku.get(sku, 0),
                "moq": int(moq_by_sku.get(sku, 0)),
                "open_demand": 0,
                "stock_gap": 0,
                "future_qty": 0,
                "uncovered_qty": 0,
            }
            for sku, qty in base_stock.items()
        },
        "future_po_by_sku": {
            sku: [
                {
                    "po": po["po"],
                    "qty": int(po["qty"]),
                    "available": fdate(po["available"]),
                    "reliability": po["reliability"],
                    "bol": po["bol"],
                    "factory_date": fdate(po.get("factory_date")),
                    "sailing_date": fdate(po.get("sailing_date")),
                    "port_eta": fdate(po.get("port_eta")),
                    "actual_receipt_date": fdate(po.get("actual_receipt_date")),
                }
                for po in pos
            ]
            for sku, pos in future_pos_by_sku.items()
        },
        "summary": {
            "open_so_count": len(so_board),
            "critical_so_count": len(critical_sos),
            "watch_so_count": len(watch_sos),
            "issue_sku_count": len([s for s in sku_summary if s["issue_count"]]),
            "action_count": len(action_queue),
            "old_open_line_count": len(old_open_lines),
            "data_check_count": len(old_open_lines) + len([x for x in purchase_rows if x["reliability"] in {"未发货-需确认工厂交期", "无可用到仓日期"} and x["qty"] > 0]) + len([
                po
                for sku, pos in future_pos_by_sku.items()
                for po in pos
                if po.get("actual_receipt_date")
            ]),
        },
        "so_board": so_board,
        "so_lines": sorted(so_lines, key=lambda x: (status_rank(x["status"]), x["latest_ship"] or "9999-12-31", x["so"], x["sku"])),
        "sku_summary": sorted(sku_summary, key=lambda x: (-x["critical_count"], -x["watch_count"], -max(x["stock_gap"], 0), x["sku"])),
        "recommendations": sorted(recommendations, key=lambda x: (-max(x["issue_qty"], x["stock_gap"]), x["sku"])),
        "po_impact": po_rows,
        "purchase_rows": purchase_rows,
        "shipment_control_rows": shipment_control_rows,
        "replenishment": sorted(replenishment_rows, key=lambda x: (0 if x["suggested_po_qty"] else 1, -x["suggested_po_qty"], x["sku"])),
        "reminders": reminders,
        "action_queue": action_queue,
        "data_checks": {
            "old_open_lines": old_open_lines[:200],
            "purchase_without_usable_eta": [x for x in purchase_rows if x["reliability"] in {"未发货-需确认工厂交期", "无可用到仓日期"} and x["qty"] > 0][:200],
            "received_po_excluded": arrived_po_rows[:200],
            "po_receipt_inconsistency": [
                {
                    "issue": "DATA INCONSISTENCY",
                    "po": po["po"],
                    "sku": sku,
                    "qty": int(po["qty"]),
                    "actual_receipt_date": fdate(po.get("actual_receipt_date")),
                    "available": fdate(po.get("available")),
                    "action": "已到货 PO 不应出现在未来在途；请检查 Purchase ETA/状态字段。",
                }
                for sku, pos in future_pos_by_sku.items()
                for po in pos
                if po.get("actual_receipt_date")
            ][:200],
            "transit_unknown_dc": transit_check_rows[:200],
            "excluded_sales_lines": excluded_sales_lines[:200],
        },
    }
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    HTML_OUT.write_text(render_html(data), encoding="utf-8")
    print(HTML_OUT)
    print(JSON_OUT)


def badge(status):
    cls = "ok" if status == "OK" else ("crit" if status == "Critical Risk" else "watch")
    return f'<span class="badge {cls}">{escape(status)}</span>'


def cell(value, cls=""):
    return f'<td class="{cls}">{escape(str(value))}</td>'


def linked_value(key, value):
    text = escape(str(value))
    if not text:
        return ""
    if key in {"so", "standard_affected_so", "recommended_so"}:
        return f'<button class="link" data-kind="so" data-value="{text}">{text}</button>'
    if key == "sku":
        return f'<button class="link" data-kind="sku" data-value="{text}">{text}</button>'
    if key == "po":
        return f'<button class="link" data-kind="po" data-value="{text}">{text}</button>'
    return text


def rows(items, cols, limit=None):
    out = []
    for item in items[: limit or len(items)]:
        tds = []
        for key, label in cols:
            value = item.get(key, "")
            if key == "status":
                tds.append(f"<td>{badge(value)}</td>")
            else:
                content = linked_value(key, value)
                tds.append(f'<td class="{"num" if isinstance(value, (int, float)) else ""}">{content}</td>')
        out.append("<tr>" + "".join(tds) + "</tr>")
    return "\n".join(out) or f'<tr><td colspan="{len(cols)}" class="empty">暂无记录</td></tr>'


def table(title, items, cols, subtitle="", limit=None):
    heads = "".join(f"<th>{escape(label)}</th>" for _, label in cols)
    return f"""
    <section class="panel">
      <div class="panel-title">
        <h2>{escape(title)}</h2>
        <p>{escape(subtitle)}</p>
      </div>
      <div class="table-wrap">
        <table>
          <thead><tr>{heads}</tr></thead>
          <tbody>{rows(items, cols, limit)}</tbody>
        </table>
      </div>
    </section>
    """


def render_html(data):
    payload = json.dumps(data, ensure_ascii=False)
    s = data["summary"]
    today_tables = table("今日优先处理", data["action_queue"], [
        ("reason", "原因"), ("status", "状态"), ("so", "SO/CI"), ("customer", "客户"), ("delivery_center", "客户仓"),
        ("latest_ship", "最晚美仓发货日"), ("days_to_ship", "剩余天数"), ("required_arrival", "客户要求到仓日"), ("line_count", "行数"),
        ("issue_skus", "问题SKU数"), ("critical_skus", "高风险SKU数"), ("watch_skus", "确认在途观察SKU数"), ("action", "处理建议")
    ], "已去重：包含所有高风险 SO，以及最晚美仓发货日在 21 天内的观察 SO。")
    html = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>SCM Risk Dashboard MVP</title>
<style>
:root {{ --ink:#182331; --muted:#5e6b78; --line:#dbe4ee; --head:#eaf3fb; --crit:#f7d8da; --watch:#fff1c9; --ok:#dff0d8; --blue:#2e638f; --bg:#f6f8fb; }}
* {{ box-sizing:border-box; }}
body {{ margin:0; font:13px/1.42 Arial, sans-serif; color:var(--ink); background:var(--bg); }}
header {{ position:sticky; top:0; z-index:3; background:#fff; border-bottom:1px solid var(--line); padding:14px 22px 10px; }}
h1 {{ margin:0 0 10px; font-size:22px; letter-spacing:0; }}
.meta {{ color:var(--muted); }}
.cards {{ display:grid; grid-template-columns:repeat(6, minmax(120px,1fr)); gap:8px; margin-top:12px; }}
.card {{ background:#f7fbff; border:1px solid var(--line); border-radius:8px; padding:10px 12px; }}
.card b {{ display:block; font-size:24px; margin-top:2px; }}
nav {{ display:flex; gap:6px; margin-top:12px; flex-wrap:wrap; }}
button {{ border:1px solid var(--line); background:#fff; border-radius:6px; padding:7px 10px; cursor:pointer; color:var(--ink); }}
button.active {{ background:var(--blue); color:#fff; border-color:var(--blue); }}
button.link {{ border:0; background:transparent; color:#1f5f91; padding:0; font:inherit; font-weight:700; text-decoration:underline; }}
main {{ padding:16px 22px 28px; }}
.tab {{ display:none; }}
.tab.active {{ display:block; }}
.panel {{ background:#fff; border:1px solid var(--line); border-radius:8px; margin-bottom:14px; overflow:hidden; }}
.panel-title {{ padding:12px 14px; background:var(--head); }}
h2 {{ margin:0; font-size:16px; }}
p {{ margin:3px 0 0; color:var(--muted); }}
.table-wrap {{ overflow:auto; max-height:64vh; position:relative; }}
table {{ width:max-content; min-width:100%; border-collapse:separate; border-spacing:0; table-layout:auto; }}
th,td {{ border-bottom:1px solid var(--line); padding:7px 10px; text-align:left; vertical-align:top; white-space:nowrap; min-width:92px; max-width:none; }}
th {{ position:sticky; top:0; background:#eef3f8; z-index:2; font-size:12px; line-height:1.2; }}
td {{ line-height:1.25; background:#fff; }}
th.sticky-key, td.sticky-key {{ position:sticky; left:0; z-index:3; background:#fff; box-shadow:1px 0 0 var(--line), 10px 0 12px rgba(255,255,255,.86); }}
th.sticky-key {{ z-index:4; background:#eef3f8; }}
td.long, th.long {{ min-width:240px; }}
.num {{ text-align:right; font-variant-numeric:tabular-nums; }}
.badge {{ display:inline-block; padding:3px 7px; border-radius:5px; font-weight:700; white-space:nowrap; }}
.crit {{ background:var(--crit); }}
.watch {{ background:var(--watch); }}
.ok {{ background:var(--ok); }}
.empty {{ text-align:center; color:var(--muted); padding:24px; }}
.search {{ margin:0 0 12px; width:320px; max-width:100%; padding:8px 10px; border:1px solid var(--line); border-radius:6px; }}
.filter-banner {{ display:none; align-items:center; gap:10px; margin:-4px 0 12px; padding:8px 10px; border:1px solid var(--line); border-radius:6px; background:#f7fbff; color:var(--muted); }}
.filter-banner b {{ color:var(--ink); }}
.inline-filter {{ display:none; align-items:center; gap:10px; padding:10px 14px; border-bottom:1px solid var(--line); background:#fff; color:var(--muted); }}
.inline-filter b {{ color:var(--ink); }}
.detail {{ display:none; background:#fff; border:2px solid #2e638f; border-radius:8px; margin:0 0 14px; overflow:hidden; }}
.detail.active {{ display:block; }}
.detail-head {{ padding:12px 14px; background:#e7f2fb; display:flex; align-items:flex-start; justify-content:space-between; gap:12px; }}
.detail-head h2 {{ margin:0 0 3px; }}
.detail-body {{ padding:0; }}
.detail-grid {{ display:grid; grid-template-columns:repeat(4, minmax(120px, 1fr)); gap:8px; padding:12px 14px; border-bottom:1px solid var(--line); }}
.mini {{ border:1px solid var(--line); border-radius:6px; padding:8px; background:#f9fbfd; }}
.mini b {{ display:block; font-size:18px; margin-top:2px; }}
.toolbox, .export-grid {{ display:flex; flex-wrap:wrap; gap:8px; padding:12px 14px; align-items:center; }}
.toolbox button, .export-grid button {{ background:#f8fbff; }}
.field {{ width:90px; border:1px solid var(--line); border-radius:5px; padding:5px 6px; }}
.note {{ width:180px; border:1px solid var(--line); border-radius:5px; padding:5px 6px; }}
.select {{ border:1px solid var(--line); border-radius:5px; padding:5px 6px; }}
.small-input {{ width:72px; border:1px solid var(--line); border-radius:5px; padding:5px 6px; }}
.transit-table th:nth-child(1), .transit-table td:nth-child(1) {{ min-width:120px; }}
.transit-table th:nth-child(2), .transit-table td:nth-child(2) {{ min-width:280px; }}
.transit-table th:nth-child(3), .transit-table td:nth-child(3) {{ min-width:70px; }}
.transit-table th:nth-child(4), .transit-table td:nth-child(4) {{ min-width:150px; }}
.transit-table th:nth-child(n+5), .transit-table td:nth-child(n+5) {{ min-width:86px; }}
.hint {{ padding:10px 14px; color:var(--muted); }}
@media (max-width:900px) {{ .cards {{ grid-template-columns:repeat(2,1fr); }} main, header {{ padding-left:12px; padding-right:12px; }} }}
</style>
</head>
<body>
<header>
  <h1>SCM 风险看板 MVP</h1>
  <div class="meta">生成日期 {escape(data["generated_at"])} · 库存基准日 {escape(data["base_date"])} · 数据来源：Follow Up；计算展示：本地看板</div>
  <div class="cards">
    <div class="card">未完成 SO<b id="cardOpenSo">{s["open_so_count"]}</b></div>
    <div class="card">高风险 SO<b id="cardCriticalSo">{s["critical_so_count"]}</b></div>
    <div class="card">观察 SO<b id="cardWatchSo">{s["watch_so_count"]}</b></div>
    <div class="card">问题 SKU<b id="cardIssueSku">{s["issue_sku_count"]}</b></div>
    <div class="card">今日待办<b id="cardActionQueue">{s["action_count"]}</b></div>
    <div class="card">数据检查<b id="cardDataChecks">{s["old_open_line_count"]}</b></div>
  </div>
  <nav>
    <button class="active" data-tab="today">今日待办</button>
    <button data-tab="so">SO 风险</button>
    <button data-tab="control">SO 明细</button>
    <button data-tab="sku">SKU 缺货</button>
    <button data-tab="logistics">入库催办</button>
    <button data-tab="needby">到仓红线</button>
    <button data-tab="replenishment">补货建议</button>
    <button data-tab="transit">运输设置</button>
    <button data-tab="workbook">上传 Follow Up</button>
    <button data-tab="sps">导入 SPS</button>
    <button data-tab="export">导出</button>
    <button data-tab="checks">数据检查</button>
  </nav>
</header>
<main>
  <input class="search" id="search" placeholder="筛选 SO / SKU / 客户 / PO / BOL">
  <div class="filter-banner" id="activeFilterBanner"><span id="activeFilterText"></span><button id="clearActiveFilter">清除筛选</button></div>
  <section class="panel" id="importNotice" style="display:none"></section>
  <section class="detail" id="detail"></section>
  <div class="tab active" id="today">{today_tables}</div>
  <div class="tab" id="so">
    {table("订单层级看板", data["so_board"], [("status","状态"),("so","SO/CI"),("customer","客户"),("delivery_center","客户仓"),("latest_ship","最晚美仓发货日"),("required_arrival","客户要求到仓日"),("suggested_customer_date","建议客户交期"),("line_count","行数"),("issue_skus","问题SKU数"),("critical_skus","高风险SKU数"),("planned_skus","预计在途SKU数"),("watch_skus","确认在途观察SKU数"),("action","处理建议")])}
  </div>
  <div class="tab" id="control">
    {table("订单 SKU 明细", data["so_lines"], [("status","状态"),("so","SO/CI"),("customer","客户"),("delivery_center","客户仓"),("sku","SKU"),("product","产品"),("qty","数量"),("risk_qty","风险数量"),("required_arrival","客户要求到仓日"),("latest_ship","最晚美仓发货日"),("suggested_customer_date","建议客户交期"),("stock_before","分配前库存"),("stock_used","使用库存"),("cover_type","覆盖类型"),("cover","覆盖来源"),("uncovered","未覆盖数量"),("action","处理建议")])}
  </div>
  <div class="tab" id="sku">
    {table("SKU 层级缺口", data["sku_summary"], [("sku","SKU"),("product","产品"),("current_onhand","当前库存"),("open_demand","未交货需求"),("stock_gap","库存缺口"),("future_qty","未来PO数量"),("net_gap_after_future","未来PO后缺口"),("issue_qty","风险数量"),("critical_qty","高风险数量"),("planned_qty","预计在途数量"),("watch_qty","确认在途观察数量"),("uncovered_qty","未覆盖数量"),("issue_count","问题行数"),("critical_count","高风险行数"),("planned_count","预计在途行数"),("watch_count","确认在途观察行数")])}
  </div>
  <div class="tab" id="logistics"></div>
  <div class="tab" id="needby"></div>
  <div class="tab" id="replenishment">
    {table("Replenishment Suggestion / 预测补货建议", data["replenishment"], [("status","状态"),("sku","SKU"),("product","产品"),("weekly_forecast","周预测"),("lead_time_weeks","LT周数"),("safety_stock_weeks","安全库存周数"),("buffer_weeks","缓冲周数"),("target_weeks","目标周数"),("target_stock","目标库存"),("total_stock","Total Stock"),("formula_gap","公式缺口"),("case_pack","箱规"),("suggested_po_qty","建议下单量"),("sum_purchase_suggestion","Sum原建议"),("order_risk_gap","订单风险缺口"),("onhand","当前库存"),("open_demand","未交货需求"),("confirmed_incoming","确认在途"),("planned_incoming","预计在途"),("moq","MOQ"),("action","说明")], "主公式：MAX(0, Weekly Forecast × (LT + SS + 2) - Total Stock)，再按箱规向上取整。订单风险缺口只做参考，不参与建议采购量。")}
  </div>
  <div class="tab" id="transit"></div>
  <div class="tab" id="workbook">
    <section class="panel">
      <div class="panel-title">
        <h2>Follow Up Upload / 上传最新版 Follow Up</h2>
        <p>Use this when order dates, purchase ETA, inventory, or sales lines changed. This requires opening the dashboard through the local server URL.</p>
      </div>
      <div class="toolbox">
        <input type="file" id="followupFile" accept=".xlsx,.xlsm">
        <button id="uploadFollowup">Upload and Recalculate</button>
      </div>
      <div id="followupUploadStatus" class="hint">如果这个页面是直接打开的文件，请先启动本地看板服务。</div>
    </section>
  </div>
  <div class="tab" id="sps">
    <section class="panel">
      <div class="panel-title">
        <h2>SPS Import / 导入 SPS 新订单</h2>
        <p>Choose one or more SPS CSV files. If SPS gives you a zip, unzip it first and select the CSV files inside.</p>
      </div>
      <div class="toolbox">
        <input type="file" id="spsFiles" multiple accept=".csv,.txt">
        <button id="confirmSpsImport">Confirm Import Preview</button>
        <button id="clearSps">Clear Import</button>
        <button id="clearConfirmedImports">Clear Confirmed Imports</button>
        <button id="downloadSpsDiff">Download SPS Difference CSV</button>
        <button id="copySpsNew">Copy New Lines for Excel</button>
        <button id="downloadSpsSalesXls">Download Sales Paste File</button>
      </div>
      <div id="spsSummary" class="detail-grid"></div>
      <div id="spsRiskPreview"></div>
      <div id="spsResults" class="hint">No SPS files imported yet.</div>
    </section>
  </div>
  <div class="tab" id="export">
    <section class="panel">
      <div class="panel-title">
        <h2>Export / 导出数据</h2>
        <p>导出 CSV 或可粘贴表，用于 Follow Up、Zoho、催工厂/货代/仓库。</p>
      </div>
      <div class="export-grid">
        <button data-export="action_queue">今日待办</button>
        <button data-export="so_board">SO 风险汇总</button>
        <button data-export="so_lines">SO-SKU 明细</button>
        <button data-export="sku_summary">SKU 缺货汇总</button>
        <button data-export="replenishment">预测补货建议</button>
        <button data-export="po_impact">PO 影响明细</button>
        <button data-export="logistics_actions">催工厂/货代/仓库</button>
        <button data-export="inbound_need_by">在途到仓红线</button>
        <button data-export="manual_audit">人工分配检查</button>
        <button data-export="warehouse_email">21天邮件提醒</button>
        <button data-export="confirmed_sps_sales_rows">SPS新单粘贴行</button>
        <button id="downloadConfirmedSpsXls">SPS新单粘贴文件</button>
        <button id="downloadAllocations">人工分配记录</button>
      </div>
      <div class="hint">交期、库存、采购变化请上传最新版 Follow Up 后重新计算。</div>
    </section>
  </div>
  <div class="tab" id="checks">
    {table("DATA INCONSISTENCY / 数据不一致", data["data_checks"]["po_receipt_inconsistency"], [("issue","问题"),("po","PO"),("sku","SKU"),("qty","数量"),("actual_receipt_date","实际到货日期"),("available","预计到仓日"),("action","处理建议")], "如果已到货 PO 仍然进入未来在途，这里会强提醒。")}
    {table("已到货 PO 排除记录", data["data_checks"]["received_po_excluded"], [("row","Purchase 行"),("po","PO"),("sku","SKU"),("product","产品"),("qty","数量"),("actual_receipt_date","实际到货日期"),("sailing_date","船期"),("port_eta","到港日"),("bol","BOL"),("status","状态")], "这些 PO 已按 Sum Onhand 体现，不再计入确认在途/预计在途。")}
    {table("期初前未完结订单", data["data_checks"]["old_open_lines"], [("row","Sales 行"),("so","SO/CI"),("sku","SKU"),("qty","数量"),("order_date","订单日期"),("required_arrival","客户要求到仓日")])}
    {table("没有可靠到仓日期的采购", data["data_checks"]["purchase_without_usable_eta"], [("row","Purchase 行"),("po","PO"),("sku","SKU"),("product","产品"),("qty","数量"),("factory_date","工厂发货时间"),("sailing_date","船期"),("port_eta","到港日"),("available_date","预计到仓日"),("reliability","状态")])}
    {table("运输天数待维护", data["data_checks"]["transit_unknown_dc"], [("customer","客户"),("delivery_center","客户仓"),("current_days","当前天数"),("issue","问题"),("action","处理建议")], "这些客户仓暂时按默认 7 天计算，请在运输设置页维护。")}
    {table("已排除订单 / Petco MX & Canada", data["data_checks"]["excluded_sales_lines"], [("row","Sales 行"),("so","SO/CI"),("sku","SKU"),("qty","数量"),("customer","客户"),("delivery_center","客户仓"),("required_arrival","客户要求到仓日"),("reason","原因")], "这些订单不参与 US 仓发货风险、补货建议和 21 天提醒。")}
  </div>
</main>
<script>
window.SCM_DATA = {payload};
function esc(v) {{
  return String(v ?? '').replace(/[&<>"']/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));
}}
function zhStatus(v) {{
  const map = {{
    'Critical Risk':'高风险',
    'Warehouse Prep Risk':'仓库备货风险',
    'Planned PO Watch':'预计在途待确认',
    'Lead Time Watch':'交期观察',
    'OK':'正常',
    'Late Risk':'逾期风险',
    'Buffer Stock':'缓冲库存',
    'SO Needed':'订单需求',
    'Forecast Needed':'预测需求',
    'Unallocated Stock':'库存未充分分配',
    'Leftover Stock':'剩余库存',
    'Sent':'已发送',
    'Not Due':'未到期',
    'Overdue':'已逾期',
    'Due Today':'今日到期',
    'Upcoming':'即将到期',
    'Confirmed':'已确认',
  }};
  return map[v] || v;
}}
function zhAction(text) {{
  let s = String(text ?? '');
  const exact = {{
    'Covered by current stock.':'现有库存可覆盖。',
    'Covered by confirmed incoming PO on time.':'确认在途可按时覆盖。',
    'Covered only if planned ETA holds; confirm sailing date before promising.':'仅在预计到仓成立时可覆盖；承诺前请确认船期/交期。',
    'Supplier lead time is tight; confirm PO timing.':'供应商交期偏紧，请确认 PO 时间。',
    'Ask sales/customer to revise date or expedite cover PO.':'请与销售/客户沟通改期，或催促覆盖 PO 提前。',
    'Confirm sailing date/factory timing for planned ETA PO.':'请确认预计在途 PO 的船期/工厂交期。',
    'Confirmed incoming PO covers; monitor ETA/receiving.':'确认在途可覆盖，跟进入仓/卸货时间。',
    'No action.':'无需处理。',
    'No action after manual allocation.':'人工分配后无需处理。',
    'Manual allocation still leaves critical SKU risk.':'人工分配后仍有高风险 SKU。',
    'Manual allocation depends on planned ETA; confirm sailing date.':'人工分配依赖预计到仓，请确认船期。',
    'Manual allocation leaves confirmed incoming PO watch.':'人工分配后仍需跟进确认在途。',
    'Need additional PO or customer date change.':'需要追加 PO 或与客户沟通改期。',
    'Warehouse prep time is short; suggest customer date change.':'仓库备货时间不足，建议沟通客户改期。',
    'Covered late; follow BOL inbound and suggest customer date if needed.':'到仓晚于最晚发货日；请跟进 BOL 清关/送仓，必要时建议客户改期。',
    'Covered late; follow PO timing or suggest customer date.':'覆盖时间偏晚；请跟进 PO 时间或建议客户改期。',
    'Covered by planned ETA; confirm factory shipment/sailing.':'预计在途可覆盖，请确认工厂发货/船期。',
    'Covered by incoming PO; monitor receiving date.':'在途 PO 可覆盖，请跟进入仓日期。',
    'Can be covered by current stock.':'现有库存可覆盖。',
    'Covered if planned ETA holds; confirm sailing date.':'预计到仓成立即可覆盖，请确认船期。',
    'Covered by incoming PO under current customer date.':'按当前客户交期，在途 PO 可覆盖。',
    'Confirmed SPS line included in local dashboard calculation.':'已确认 SPS 新单已纳入本地看板计算。',
    'New SPS confirmed import':'已确认 SPS 新单',
    'Included in local dashboard':'已纳入本地看板',
    'Warehouse email marked sent.':'仓库邮件已标记发送。',
    'Missing required arrival date.':'缺少客户要求到仓日。',
    'Send warehouse notice and mark sent.':'发送仓库通知并标记已发送。',
    'Not due yet.':'尚未到提醒日期。',
    'Current stock covers all demand.':'现有库存已覆盖全部需求。',
    'Covered after known POs':'已知 PO 到仓后可覆盖',
    'Urgent: push customs clearance, truck appointment, and delivery to WH.':'紧急：催清关、卡车预约和送仓。',
    'Confirm clearance, appointment, and delivery date.':'确认清关、预约和送仓日期。',
    'Ask warehouse to prioritize receiving/release immediately after delivery.':'请仓库到货后优先卸货/释放库存。',
    'Ask warehouse to plan receiving before latest ship date.':'请仓库安排在最晚发货日前完成收货。',
    'No urgent follow-up based on current SO and forecast.':'按当前 SO 和预测暂无紧急跟进。',
    'Monitor BOL receiving and release before latest ship date.':'跟进 BOL 到仓入库，确保最晚发货日前释放库存。',
    'Very tight: prioritize clearance, delivery, receiving, and release.':'时间非常紧：优先催清关、送仓、卸货和释放库存。',
    'Confirm factory shipment, sailing date, and BOL.':'确认工厂发货、船期和 BOL。',
    'Critical risk':'高风险',
    'Overdue critical risk':'已逾期高风险',
    'Planned ETA needs confirmation':'预计到仓需确认',
    '21-day warehouse notice':'提前 21 天通知仓库',
    '仓库备货风险':'仓库备货风险',
    '已到港但 Purchase 还没有到仓 ATD，请催货代/仓库清关送仓；上传 Follow Up 后如果 Purchase ATD 有值，这条提醒会消失。':'已到港但 Purchase 还没有到仓 ATD，请催货代/仓库清关送仓；上传 Follow Up 后如果 Purchase ATD 有值，这条提醒会消失。',
    'Sales 仍未填写 ATD，请确认美国仓是否已安排发货/配送；上传 Follow Up 后如果 Sales ATD 有值，这条提醒会消失。':'Sales 仍未填写 ATD，请确认美国仓是否已安排发货/配送；上传 Follow Up 后如果 Sales ATD 有值，这条提醒会消失。',
  }};
  if (exact[s]) return exact[s];
  s = s.replace(/^Short (\\d+) units\\. Need new\\/earlier PO\\.$/, '缺 $1 件，需要新增或提前 PO。');
  s = s.replace(/^Short (\\d+) units after current stock and known POs\\.$/, '现有库存和已知 PO 后仍缺 $1 件。');
  s = s.replace(/^Need earlier ETA for (.+) by (\\d+) days\\.$/, '需要 $1 提前 $2 天到仓。');
  s = s.replace(/^Warehouse needs 7 days after arrival; suggest customer date ([0-9-]+)\\.$/, '仓库到仓后需要 7 天备货；建议客户交期改为 $1。');
  s = s.replace(/^Manual allocation still leaves critical SKU risk; suggest customer date ([0-9-]+)\\.$/, '人工分配后仍有高风险 SKU；建议客户交期改为 $1。');
  s = s.replace(/^Need earlier ETA for (.+) by (\\d+) days\\. Cover is planned ETA, confirm sailing date\\.$/, '需要 $1 提前 $2 天到仓；该覆盖为预计在途，请确认船期。');
  s = s.replace(/^Covered late by (\\d+) days\\. Need earlier PO\\/ship plan\\.$/, '覆盖晚 $1 天，需要提前 PO/船期安排。');
  s = s.replace(/^Covered late by (\\d+) days\\. Need earlier PO\\/ship plan\\. Cover is planned ETA; confirm sailing date\\.$/, '覆盖晚 $1 天，需要提前 PO/船期安排；该覆盖为预计在途，请确认船期。');
  s = s.replace(/ Suggest customer date ([0-9-]+)\\.$/, '；建议客户交期改为 $1。');
  s = s.replace(/^Total demand is short by (\\d+) after all known POs; consider add-on PO or customer date change\\.$/, '所有已知 PO 后仍缺 $1 件；考虑加单或沟通客户改期。');
  s = s.replace(/^Current stock \\+ confirmed incoming is short by (\\d+); coverage depends on planned PO or timing\\.$/, '现有库存 + 确认在途仍缺 $1 件；覆盖依赖预计 PO 或时间变化。');
  s = s.replace(/^Total demand is covered by current stock and confirmed incoming; watch line timing if any order is flagged\\.$/, '总需求可由现有库存和确认在途覆盖；如订单行仍预警，请关注时间节点。');
  s = s.replace(/^Need (\\d+) more units from incoming PO or add-on PO\\.$/, '还需要 $1 件在途或追加 PO 覆盖。');
  s = s.replace(/^ETA is late by (\\d+) days versus need-by date\\.$/, '预计到仓比需求红线晚 $1 天。');
  s = s.replace(/^Late by (\\d+) days; communicate customer date or expedite receiving\\.$/, '晚 $1 天；请沟通客户改期，或加急清关/送仓/卸货。');
  s = s.replace(/^Need PO\\/factory timing earlier by (\\d+) days, or communicate customer date\\.$/, '需要 PO/工厂时间提前 $1 天，或沟通客户改期。');
  s = s.replace(/^Need planned PO to hold; reliable supply is short by (\\d+)\\.$/, '需要预计 PO 按计划执行；可靠供应仍缺 $1 件。');
  s = s.replace(/^Still short (\\d+) after all known POs; consider add-on PO or customer date change\\.$/, '所有已知 PO 后仍缺 $1 件；考虑加单或沟通客户改期。');
  s = s.replace(/^Need by ([0-9-]+) based on open SO\\.$/, '按未交货 SO，最晚需要 $1 到仓。');
  s = s.replace(/^Need by ([0-9-]+) based on forecast\\.$/, '按预测需求，最晚需要 $1 到仓。');
  s = s.replace(/^First consumed by SO (.+)\\.$/, '最早被 SO $1 消耗。');
  return s;
}}
function statusBadge(v) {{
  const cls = v === 'OK' || v === 'Buffer Stock' ? 'ok' : (v === 'Critical Risk' || v === 'Late Risk' ? 'crit' : 'watch');
  return `<span class="badge ${{cls}}">${{esc(zhStatus(v))}}</span>`;
}}
function isPlannedReliability(value) {{
  return ['Planned ETA','预计在途'].includes(String(value || ''));
}}
function isReceivedPo(po) {{
  return !!(po?.actual_receipt_date || po?.received_date || po?.receipt_date);
}}
function usableFuturePos(sku) {{
  return (window.SCM_DATA.future_po_by_sku?.[sku] || []).filter(po => !isReceivedPo(po));
}}
function isPlannedCover(covers) {{
  return covers.some(c => isPlannedReliability(c.reliability));
}}
function statusRank(status) {{
  return status === 'Critical Risk' ? 0 : status === 'Warehouse Prep Risk' ? 1 : status === 'Planned PO Watch' ? 2 : status === 'Lead Time Watch' ? 3 : 4;
}}
function mini(label, value) {{ return `<div class="mini">${{esc(label)}}<b>${{esc(value)}}</b></div>`; }}
function skuInfoFor(sku) {{
  return window.SCM_DATA.sku_summary.find(x => x.sku === sku)
    || window.SCM_DATA.inventory_by_sku?.[sku]
    || {{sku, product:window.SCM_DATA.product_name_by_sku?.[sku] || '', current_onhand:0, weekly_forecast:0, open_demand:0, stock_gap:0, future_qty:0, uncovered_qty:0}};
}}
function simpleTable(items, cols) {{
  const longCols = new Set(['product','action','forwarder_action','warehouse_action','delivery_center','related_pos','urgent_skus','affected_sos','reason']);
  const keyCols = new Set(['so','order','po','sku','bol']);
  const stickyKey = (cols.find(c => keyCols.has(c[0])) || [null])[0];
  const cellClass = (field, value='') => `${{typeof value === 'number' ? 'num ' : ''}}${{longCols.has(field) ? 'long ' : ''}}${{field === stickyKey ? 'sticky-key' : ''}}`.trim();
  const head = cols.map(c => `<th class="${{cellClass(c[0])}}">${{esc(c[1])}}</th>`).join('');
  const body = items.length ? items.map(item => `<tr>${{cols.map(c => {{
    const v = item[c[0]] ?? '';
    if (c[0] === 'status') return `<td>${{statusBadge(v)}}</td>`;
    if (['action','reason','conclusion'].includes(c[0])) return `<td class="${{cellClass(c[0], v)}}" title="${{esc(zhAction(v))}}">${{esc(zhAction(v))}}</td>`;
    if (['so','sku','po'].includes(c[0])) return `<td class="${{cellClass(c[0], v)}}"><button class="link" data-kind="${{c[0]}}" data-value="${{esc(v)}}">${{esc(v)}}</button></td>`;
    if (c[0] === 'bol' && v) return `<td class="${{cellClass(c[0], v)}}"><button class="link" data-kind="bol" data-value="${{esc(v)}}">${{esc(v)}}</button></td>`;
    return `<td class="${{cellClass(c[0], v)}}" title="${{esc(v)}}">${{esc(v)}}</td>`;
  }}).join('')}}</tr>`).join('') : `<tr><td class="empty" colspan="${{cols.length}}">暂无记录</td></tr>`;
  return `<div class="table-wrap"><table><thead><tr>${{head}}</tr></thead><tbody>${{body}}</tbody></table></div>`;
}}
function allocationKey(sku, so) {{ return `${{sku}}__${{so}}`; }}
function getAllocations() {{
  try {{ return JSON.parse(localStorage.getItem('scm_manual_allocations') || '{{}}'); }}
  catch {{ return {{}}; }}
}}
function saveAllocation(sku, so, field, value) {{
  const all = getAllocations();
  const key = allocationKey(sku, so);
  all[key] = all[key] || {{sku, so}};
  all[key][field] = value;
  if (field === 'assign_qty') {{
    all[key].assign = String(value).trim() === '' ? 'system' : 'manual';
  }}
  if (field === 'assign' && value === 'system') {{
    all[key].assign_qty = '';
  }}
  all[key].updated_at = new Date().toISOString();
  localStorage.setItem('scm_manual_allocations', JSON.stringify(all));
}}
function isDecisionRow(x) {{
  return !!x && ((x.assign && x.assign !== 'system') || String(x.assign_qty ?? '').trim() !== '' || String(x.note || '').trim());
}}
function updateAllocationSummary(sku) {{
  const box = document.getElementById('allocSummary');
  if (!box) return;
  const lines = window.SCM_DATA.so_lines.filter(x => x.sku === sku);
  const {{manualCalc, savedRows, stockLimit}} = manualCalcForSku(sku, lines);
  const assignedQty = [...manualCalc.values()].reduce((sum, value) => sum + value.manualQty, 0);
  const remainingQty = Math.max(stockLimit - assignedQty, 0);
  box.innerHTML = mini('当前库存', stockLimit) + mini('人工分配数量', assignedQty) + mini('剩余库存', remainingQty) + mini('决策行数', savedRows.length);
}}
function capManualStockInput(el) {{
  const sku = el.dataset.sku;
  const so = el.dataset.so;
  if (String(el.value ?? '').trim() === '') return '';
  const skuInfo = skuInfoFor(sku);
  const stockLimit = Number(skuInfo.current_onhand || 0);
  const skuLines = window.SCM_DATA.so_lines.filter(x => x.sku === sku);
  const line = skuLines.find(x => x.so === so) || {{}};
  const rowLimit = Number(line.qty || 0);
  const all = getAllocations();
  const otherAssigned = skuLines
    .filter(x => x.so !== so)
    .reduce((sum, x) => {{
      const saved = all[allocationKey(sku, x.so)] || {{}};
      const hasManualQty = String(saved.assign_qty ?? '').trim() !== '';
      return sum + (hasManualQty ? Number(saved.assign_qty || 0) : Number(x.stock_used || 0));
    }}, 0);
  const maxAllowed = Math.max(Math.min(rowLimit, stockLimit - otherAssigned), 0);
  let value = Number(el.value || 0);
  if (!Number.isFinite(value) || value < 0) value = 0;
  if (value > maxAllowed) {{
    value = maxAllowed;
    el.value = value || '';
    el.title = `Max available for this row is ${{maxAllowed}}`;
    alert(`This SKU only has ${{stockLimit}} current on-hand. Max available for this row is ${{maxAllowed}}.`);
  }}
  return value;
}}
function manualStatus(line, manualQty, manualMode) {{
  if (!manualMode) return {{status:line.status, gap:line.risk_qty || 0, note:'System allocation.'}};
  const qty = Number(line.qty || 0);
  const riskQty = Math.max(qty - Number(manualQty || 0), 0);
  if (riskQty <= 0) return {{status:'OK', gap:0, note:'Covered by manual stock allocation.'}};
  if (line.cover && line.cover !== 'Current stock' && line.cover !== '现有库存') return {{status:line.status === 'Critical Risk' || line.status === 'Warehouse Prep Risk' ? line.status : (isPlannedReliability(line.cover_type) ? 'Planned PO Watch' : 'Lead Time Watch'), gap:riskQty, note:line.status === 'Warehouse Prep Risk' ? `仓库备货时间不足；建议客户交期 ${{line.suggested_customer_date || ''}}。` : `${{riskQty}} 件需要 ${{isPlannedReliability(line.cover_type) ? '确认预计在途' : 'PO/在途覆盖'}}。`}};
  return {{status:'Critical Risk', gap:riskQty, note:`${{riskQty}} units not covered by manual stock.`}};
}}
function manualCalcForSku(sku, lines) {{
  const all = getAllocations();
  const skuInfo = skuInfoFor(sku);
  const savedRows = lines.map(line => all[allocationKey(sku, line.so)] || {{}}).filter(isDecisionRow);
  const manualMode = savedRows.some(x => String(x.assign_qty ?? '').trim() !== '');
  const stockLimit = Number(skuInfo.current_onhand || 0);
  let rollingStock = stockLimit;
  const manualCalc = new Map();
  for (const line of lines) {{
    const saved = all[allocationKey(sku, line.so)] || {{}};
    const hasManualQty = String(saved.assign_qty ?? '').trim() !== '';
    let manualQty = hasManualQty ? Number(saved.assign_qty || 0) : Number(line.stock_used || 0);
    manualQty = Math.max(Math.min(manualQty, Number(line.qty || 0), rollingStock), 0);
    rollingStock -= manualQty;
    const after = manualStatus(line, manualQty, manualMode);
    manualCalc.set(line.so, {{manualQty, manualMode, after}});
  }}
  return {{manualCalc, savedRows, stockLimit, manualMode}};
}}
function allocationTable(sku, lines) {{
  const all = getAllocations();
  const {{manualCalc, savedRows, stockLimit, manualMode}} = manualCalcForSku(sku, lines);
  const assignedQty = [...manualCalc.values()].reduce((sum, value) => sum + value.manualQty, 0);
  const remainingQty = Math.max(stockLimit - assignedQty, 0);
  const body = lines.map(line => {{
    const saved = all[allocationKey(sku, line.so)] || {{}};
    const hasManualQty = String(saved.assign_qty ?? '').trim() !== '';
    const calc = manualCalc.get(line.so) || {{manualQty:0, after:manualStatus(line, 0, manualMode)}};
    const manualQty = calc.manualQty;
    const after = calc.after;
    const checked = (saved.assign === 'manual' || hasManualQty) ? 'selected' : '';
    const hold = saved.assign === 'hold' ? 'selected' : '';
    const system = (!saved.assign || saved.assign === 'system') && !hasManualQty ? 'selected' : '';
    return `<tr>
      <td>${{statusBadge(line.status)}}</td>
      <td><button class="link" data-kind="so" data-value="${{esc(line.so)}}">${{esc(line.so)}}</button></td>
      <td>${{esc(line.delivery_center)}}</td>
      <td class="num">${{esc(line.qty)}}</td>
      <td class="num">${{esc(line.risk_qty)}}</td>
      <td class="num">${{esc(line.stock_used || 0)}}</td>
      <td>${{esc(line.latest_ship)}}</td>
      <td>${{esc(line.required_arrival)}}</td>
      <td>${{esc(line.cover)}}</td>
      <td>${{statusBadge(after.status)}}</td>
      <td class="num">${{esc(after.gap)}}</td>
      <td>
        <select class="select alloc-input" data-sku="${{esc(sku)}}" data-so="${{esc(line.so)}}" data-field="assign">
          <option value="system" ${{system}}>Follow system</option>
          <option value="manual" ${{checked}}>Manual override</option>
          <option value="hold" ${{hold}}>Hold / review</option>
        </select>
      </td>
      <td><input class="field alloc-input" type="number" min="0" max="${{esc(line.qty)}}" data-sku="${{esc(sku)}}" data-so="${{esc(line.so)}}" data-field="assign_qty" value="${{esc(saved.assign_qty ?? '')}}" placeholder="${{esc(line.stock_used || 0)}}"></td>
      <td><input class="note alloc-input" data-sku="${{esc(sku)}}" data-so="${{esc(line.so)}}" data-field="note" value="${{esc(saved.note ?? '')}}" placeholder="Note"></td>
    </tr>`;
  }}).join('');
  return `<section class="panel"><div class="panel-title"><h2>Manual Stock Allocation / 人工库存分配</h2><p>Blank means follow system allocation. Enter 0 to explicitly remove stock from that SO. Total manual stock allocation cannot exceed current on-hand.</p><p><button class="clear-sku-allocation" data-sku="${{esc(sku)}}">Clear this SKU manual allocation</button></p></div>
    <div class="detail-grid" id="allocSummary">${{mini('当前库存', stockLimit)}}${{mini('人工分配数量', assignedQty)}}${{mini('剩余库存', remainingQty)}}${{mini('决策行数', savedRows.length)}}</div>
    <div class="table-wrap"><table><thead><tr><th>系统状态</th><th>SO/CI</th><th>客户仓</th><th>数量</th><th>风险数量</th><th>系统分配库存</th><th>最晚发货日</th><th>客户要求到仓日</th><th>系统覆盖来源</th><th>人工后状态</th><th>人工后缺口</th><th>决策</th><th>人工库存数量</th><th>备注</th></tr></thead><tbody>${{body}}</tbody></table></div></section>`;
}}
function systemAllocateLines(sourceLines) {{
  const bySku = new Map();
  for (const line of sourceLines) {{
    if (!bySku.has(line.sku)) bySku.set(line.sku, []);
    bySku.get(line.sku).push(line);
  }}
  const out = [];
  for (const [sku, skuLines] of bySku.entries()) {{
    const skuInfo = skuInfoFor(sku);
    let stock = Number(skuInfo.current_onhand || 0);
    const pos = usableFuturePos(sku).map(x => ({{...x, remaining:Number(x.qty || 0)}})).sort((a,b) => String(a.available).localeCompare(String(b.available)) || String(a.po).localeCompare(String(b.po)));
    const ordered = skuLines.slice().sort((a,b) => String(a.latest_ship || '9999-12-31').localeCompare(String(b.latest_ship || '9999-12-31')) || String(a.required_arrival || '9999-12-31').localeCompare(String(b.required_arrival || '9999-12-31')) || String(a.so).localeCompare(String(b.so)));
    for (const line of ordered) {{
      let need = Number(line.qty || 0);
      const stockBefore = Math.max(stock, 0);
      const stockUsed = Math.min(stockBefore, need);
      stock -= stockUsed;
      need -= stockUsed;
      const covers = [];
      for (const po of pos) {{
        if (need <= 0) break;
        if (po.remaining <= 0) continue;
        const take = Math.min(po.remaining, need);
        po.remaining -= take;
        need -= take;
        covers.push({{po:po.po, qty:take, available:po.available, reliability:po.reliability, bol:po.bol || '', factory_date:po.factory_date || '', sailing_date:po.sailing_date || '', port_eta:po.port_eta || ''}});
      }}
      const coverDate = covers.map(x => x.available).sort().slice(-1)[0] || '';
      const warehouseReady = warehouseReadyDate(coverDate);
      const plannedCover = isPlannedCover(covers);
      let status = 'OK';
      let action = 'Covered by current stock.';
      if (need > 0) {{
        status = 'Critical Risk';
        action = `Short ${{need}} units after current stock and known POs.`;
      }} else if (coverDate && line.latest_ship && coverDate > line.latest_ship) {{
        status = 'Critical Risk';
        action = `Covered late by ${{diffDays(coverDate, line.latest_ship)}} days. Need earlier PO/ship plan.${{plannedCover ? ' Cover is planned ETA; confirm sailing date.' : ''}}`;
      }} else if (warehouseReady && line.latest_ship && warehouseReady > line.latest_ship) {{
        status = 'Warehouse Prep Risk';
        action = `Warehouse needs 7 days after arrival; suggest customer date ${{suggestedCustomerDate(line, warehouseReady)}}.`;
      }} else if (covers.length) {{
        status = plannedCover ? 'Planned PO Watch' : 'OK';
        action = plannedCover ? 'Covered only if planned ETA holds; confirm sailing date before promising.' : 'Covered by confirmed incoming PO on time.';
      }}
      const cover = covers.length ? covers.map(c => `${{c.po}}: ${{c.qty}} @ ${{c.available}} (${{c.reliability}})`).join('; ') : 'Current stock';
      const suggestedDate = (status === 'Warehouse Prep Risk' && warehouseReady) ? suggestedCustomerDate(line, warehouseReady) : (coverDate && line.latest_ship && coverDate > line.latest_ship ? suggestedCustomerDate(line, coverDate) : '');
      if (status === 'Critical Risk' && suggestedDate && line.required_arrival && suggestedDate <= line.required_arrival) {{
        status = plannedCover ? 'Planned PO Watch' : 'OK';
        action = plannedCover ? 'Covered if planned ETA holds; confirm sailing date.' : 'Covered by incoming PO under current customer date.';
      }} else if (status === 'Critical Risk' && suggestedDate) {{
        action += ` Suggest customer date ${{suggestedDate}}.`;
      }}
      out.push({{
        ...line,
        status,
        system_status: status,
        stock_before: stockBefore,
        stock_used: stockUsed,
        cover,
        cover_details: covers,
        cover_type: plannedCover ? '预计在途' : (covers.length ? '确认在途' : '现有库存'),
        cover_eta: coverDate,
        warehouse_ready_date: warehouseReady,
        suggested_customer_date: suggestedDate,
        uncovered: need,
        risk_qty: status === 'OK' ? 0 : Math.max(Number(line.qty || 0) - stockUsed, need),
        action,
      }});
    }}
  }}
  return out;
}}
function currentViewData() {{
  const grouped = new Map();
  const adjustedSales = window.SCM_DATA.so_lines.map(withTransitDates);
  const allSourceLines = systemAllocateLines([...adjustedSales, ...confirmedSpsAsSalesLines().map(withTransitDates)]);
  for (const line of allSourceLines) {{
    if (!grouped.has(line.sku)) grouped.set(line.sku, []);
    grouped.get(line.sku).push(line);
  }}
  const lines = [];
  for (const [sku, skuLines] of grouped.entries()) {{
    const {{manualCalc, manualMode}} = manualCalcForSku(sku, skuLines);
    for (const line of skuLines) {{
      const calc = manualCalc.get(line.so);
      if (manualMode && calc) {{
        lines.push({{
          ...line,
          system_status: line.status,
          system_risk_qty: line.risk_qty,
          status: calc.after.status,
          risk_qty: calc.after.gap,
          manual_stock_qty: calc.manualQty,
          manual_gap: calc.after.gap,
          cover: calc.manualQty ? `Manual stock: ${{calc.manualQty}}; ${{line.cover}}` : line.cover,
          action: calc.after.note,
        }});
      }} else {{
        lines.push({{...line, system_status: line.status, system_risk_qty: line.risk_qty, manual_stock_qty:'', manual_gap:''}});
      }}
    }}
  }}
  const soMap = new Map();
  for (const line of lines) {{
    const so = soMap.get(line.so) || {{so:line.so, customer:line.customer, delivery_center:line.delivery_center, required_arrival:line.required_arrival, latest_ship:line.latest_ship, line_count:0, issue_skus:0, critical_skus:0, warehouse_skus:0, watch_skus:0, planned_skus:0, confirmed_po_skus:0, skus:[]}};
    so.required_arrival = [so.required_arrival, line.required_arrival].filter(Boolean).sort()[0] || '';
    so.latest_ship = [so.latest_ship, line.latest_ship].filter(Boolean).sort()[0] || '';
    so.line_count += 1;
    so.skus.push(line);
    soMap.set(line.so, so);
  }}
  const soBoard = [...soMap.values()].map(so => {{
    const issue = so.skus.filter(x => x.status !== 'OK');
    so.issue_skus = new Set(issue.map(x => x.sku)).size;
    so.critical_skus = new Set(issue.filter(x => x.status === 'Critical Risk').map(x => x.sku)).size;
    so.warehouse_skus = new Set(issue.filter(x => x.status === 'Warehouse Prep Risk').map(x => x.sku)).size;
    so.watch_skus = new Set(issue.filter(x => x.status === 'Lead Time Watch').map(x => x.sku)).size;
    so.planned_skus = new Set(issue.filter(x => x.status === 'Planned PO Watch').map(x => x.sku)).size;
    so.confirmed_po_skus = new Set(issue.filter(x => ['Confirmed Incoming','确认在途'].includes(x.cover_type)).map(x => x.sku)).size;
    const suggestedDates = issue.map(x => x.suggested_customer_date).filter(Boolean).sort();
    so.suggested_customer_date = suggestedDates.slice(-1)[0] || '';
    so.status = so.critical_skus ? 'Critical Risk' : (so.warehouse_skus ? 'Warehouse Prep Risk' : (so.planned_skus ? 'Planned PO Watch' : (so.watch_skus ? 'Lead Time Watch' : 'OK')));
    so.action = so.critical_skus ? (so.suggested_customer_date ? `Manual allocation still leaves critical SKU risk; suggest customer date ${{so.suggested_customer_date}}.` : 'Manual allocation still leaves critical SKU risk.') : (so.warehouse_skus ? `Warehouse needs 7 days after arrival; suggest customer date ${{so.suggested_customer_date}}.` : (so.planned_skus ? 'Manual allocation depends on planned ETA; confirm sailing date.' : (so.watch_skus ? 'Manual allocation leaves confirmed incoming PO watch.' : 'No action after manual allocation.')));
    return so;
  }}).sort((a,b) => statusRank(a.status) - statusRank(b.status) || String(a.latest_ship).localeCompare(String(b.latest_ship)) || String(a.so).localeCompare(String(b.so)));
  const actionQueue = soBoard.filter(so => {{
    const days = diffDays(so.latest_ship, window.SCM_DATA.generated_at);
    return so.status === 'Critical Risk' || (['Warehouse Prep Risk','Lead Time Watch','Planned PO Watch'].includes(so.status) && days !== '' && days <= 21);
  }}).map(so => ({{
    reason: so.status === 'Critical Risk' ? (diffDays(so.latest_ship, window.SCM_DATA.generated_at) < 0 ? 'Overdue critical risk' : 'Critical risk') : (so.status === 'Warehouse Prep Risk' ? '仓库备货风险' : (so.status === 'Planned PO Watch' ? 'Planned ETA needs confirmation' : '21-day warehouse notice')),
    ...so,
    days_to_ship: diffDays(so.latest_ship, window.SCM_DATA.generated_at),
  }}));
  const audit = [];
  for (const [sku, skuLines] of grouped.entries()) {{
    const {{manualCalc, savedRows, stockLimit, manualMode}} = manualCalcForSku(sku, skuLines);
    if (!manualMode && !savedRows.length) continue;
    const allocated = [...manualCalc.values()].reduce((sum, x) => sum + Number(x.manualQty || 0), 0);
    const remainingStock = Math.max(stockLimit - allocated, 0);
    const manualLines = skuLines.map(line => {{
      const calc = manualCalc.get(line.so) || {{manualQty:0, after:manualStatus(line,0,manualMode)}};
      return {{line, calc}};
    }});
    const remainingRiskQty = manualLines.reduce((sum, x) => sum + Number(x.calc.after.gap || 0), 0);
    const usableTargets = manualLines
      .filter(x => Number(x.calc.after.gap || 0) > 0 && remainingStock > 0)
      .map(x => `${{x.line.so}}(${{Math.min(Number(x.calc.after.gap || 0), remainingStock)}})`)
      .slice(0, 8)
      .join('; ');
    let status = 'OK';
    let action = 'No leftover current stock after manual allocation.';
    if (remainingStock > 0 && remainingRiskQty > 0) {{
      status = 'Unallocated Stock';
      action = 'Remaining current stock can still reduce risk orders.';
    }} else if (remainingStock > 0) {{
      status = 'Leftover Stock';
      action = 'Remaining current stock exists, but no open risk remains for this SKU.';
    }}
    audit.push({{
      status,
      sku,
      product:(skuInfoFor(sku)).product || '',
      current_onhand: stockLimit,
      manual_allocated: allocated,
      remaining_stock: remainingStock,
      remaining_risk_qty: remainingRiskQty,
      decision_rows: savedRows.length,
      possible_targets: usableTargets,
      action,
    }});
  }}
  audit.sort((a,b) => statusRank(a.status === 'Unallocated Stock' ? 'Critical Risk' : a.status === 'Leftover Stock' ? 'Lead Time Watch' : 'OK') - statusRank(b.status === 'Unallocated Stock' ? 'Critical Risk' : b.status === 'Leftover Stock' ? 'Lead Time Watch' : 'OK') || String(a.sku).localeCompare(String(b.sku)));
  return {{lines, soBoard, actionQueue, audit}};
}}
function buildSkuShortageRows(view) {{
  const rows = new Map();
  for (const sku of Object.keys(window.SCM_DATA.inventory_by_sku || {{}})) {{
    const base = skuInfoFor(sku);
    rows.set(sku, {{
      status:'OK',
      sku,
      product:base.product || '',
      current_onhand:Number(base.current_onhand || 0),
      open_demand:0,
      confirmed_incoming_qty:0,
      planned_incoming_qty:0,
      shortage_qty:0,
      issue_qty:0,
      affected_so_count:0,
      affected_sos:new Set(),
      follow_po_parts:new Map(),
      follow_po:'',
      follow_bol:'',
      port_eta:'',
      cover_eta:'',
      warehouse_ready_date:'',
      latest_wh_arrival:'',
      gap_days:'',
      action:'No shortage or watch.',
    }});
  }}
  for (const line of view.lines) {{
    const base = skuInfoFor(line.sku);
    const row = rows.get(line.sku) || {{
      status:'OK',
      sku:line.sku,
      product:line.product || base.product || '',
      current_onhand:Number(base.current_onhand || 0),
      open_demand:0,
      confirmed_incoming_qty:0,
      planned_incoming_qty:0,
      shortage_qty:0,
      issue_qty:0,
      affected_so_count:0,
      affected_sos:new Set(),
      follow_po_parts:new Map(),
      follow_po:'',
      follow_bol:'',
      port_eta:'',
      cover_eta:'',
      warehouse_ready_date:'',
      latest_wh_arrival:'',
      gap_days:'',
      action:'No shortage or watch.',
    }};
    row.open_demand += Number(line.qty || 0);
    if (line.status !== 'OK') {{
      row.affected_sos.add(line.so);
      row.issue_qty += Number(line.risk_qty || 0);
      if (statusRank(line.status) < statusRank(row.status)) row.status = line.status;
      if (!row.latest_wh_arrival || (line.latest_ship && line.latest_ship < row.latest_wh_arrival)) row.latest_wh_arrival = line.latest_ship || row.latest_wh_arrival;
      for (const c of (line.cover_details || [])) {{
        if (!row.cover_eta || (c.available && c.available < row.cover_eta)) row.cover_eta = c.available;
        if (!row.port_eta || (c.port_eta && c.port_eta < row.port_eta)) row.port_eta = c.port_eta;
        if (c.po) row.follow_po_parts.set(c.po, (row.follow_po_parts.get(c.po) || 0) + Number(c.qty || 0));
        if (!row.follow_bol && c.bol) row.follow_bol = c.bol;
      }}
    }}
    rows.set(line.sku, row);
  }}
  for (const [sku, row] of rows.entries()) {{
    for (const po of usableFuturePos(sku)) {{
      if (isPlannedReliability(po.reliability)) row.planned_incoming_qty += Number(po.qty || 0);
      else row.confirmed_incoming_qty += Number(po.qty || 0);
      if (!row.follow_po_parts.size && po.po) row.follow_po_parts.set(po.po, Number(po.qty || 0));
      if (!row.follow_bol && po.bol) row.follow_bol = po.bol;
      if (!row.port_eta || (po.port_eta && po.port_eta < row.port_eta)) row.port_eta = po.port_eta;
      if (!row.cover_eta || (po.available && po.available < row.cover_eta)) row.cover_eta = po.available;
    }}
    row.warehouse_ready_date = warehouseReadyDate(row.cover_eta);
    row.shortage_qty = Math.max(row.open_demand - row.current_onhand - row.confirmed_incoming_qty - row.planned_incoming_qty, 0);
    row.follow_po = [...row.follow_po_parts.entries()].map(([po, qty]) => `${{po}} - 数量 ${{Math.round(qty)}}`).join('; ');
    delete row.follow_po_parts;
    row.affected_so_count = row.affected_sos.size;
    row.affected_sos = [...row.affected_sos].slice(0,12).join('; ');
    row.gap_days = row.warehouse_ready_date && row.latest_wh_arrival ? diffDays(row.warehouse_ready_date, row.latest_wh_arrival) : (row.cover_eta && row.latest_wh_arrival ? diffDays(row.cover_eta, row.latest_wh_arrival) : '');
    if (row.shortage_qty > 0) row.action = 'Need additional PO or customer date change.';
    else if (row.status === 'Critical Risk') row.action = row.follow_bol ? 'Covered late; follow BOL inbound and suggest customer date if needed.' : 'Covered late; follow PO timing or suggest customer date.';
    else if (row.status === 'Warehouse Prep Risk') row.action = 'Warehouse prep time is short; suggest customer date change.';
    else if (row.status === 'Planned PO Watch') row.action = 'Covered by planned ETA; confirm factory shipment/sailing.';
    else if (row.status === 'Lead Time Watch') row.action = 'Covered by incoming PO; monitor receiving date.';
  }}
  return [...rows.values()]
    .filter(x => x.open_demand > 0 || x.confirmed_incoming_qty > 0 || x.planned_incoming_qty > 0 || x.current_onhand > 0)
    .sort((a,b) => statusRank(a.status)-statusRank(b.status) || Number(b.shortage_qty)-Number(a.shortage_qty) || String(a.latest_wh_arrival || '9999-12-31').localeCompare(String(b.latest_wh_arrival || '9999-12-31')) || String(a.sku).localeCompare(String(b.sku)));
}}
function buildSkuCoverageOverview(view) {{
  const bySku = new Map();
  for (const line of view.lines) {{
    if (!bySku.has(line.sku)) bySku.set(line.sku, []);
    bySku.get(line.sku).push(line);
  }}
  const rows = [];
  for (const [sku, lines] of bySku.entries()) {{
    const base = skuInfoFor(sku);
    const onhand = Number(base.current_onhand || 0);
    const totalDemand = lines.reduce((sum, x) => sum + Number(x.qty || 0), 0);
    const pos = usableFuturePos(sku).map(x => ({{...x, qty:Number(x.qty || 0)}})).sort((a,b) => String(a.available || '9999-12-31').localeCompare(String(b.available || '9999-12-31')) || String(a.po).localeCompare(String(b.po)));
    const confirmedIncoming = pos.filter(x => !isPlannedReliability(x.reliability)).reduce((sum, x) => sum + x.qty, 0);
    const plannedIncoming = pos.filter(x => isPlannedReliability(x.reliability)).reduce((sum, x) => sum + x.qty, 0);
    const reliableGap = onhand + confirmedIncoming - totalDemand;
    const totalGap = onhand + confirmedIncoming + plannedIncoming - totalDemand;
    const issueLines = lines.filter(x => x.status !== 'OK');
    if (!issueLines.length && reliableGap >= 0) continue;
    let cumulative = onhand;
    let firstCover = onhand >= totalDemand ? 'Current stock' : '';
    for (const po of pos) {{
      cumulative += po.qty;
      if (!firstCover && cumulative >= totalDemand) firstCover = `${{po.po || '(blank PO)'}} @ ${{po.available || ''}} (${{po.reliability || ''}})`;
    }}
    const status = totalGap < 0 ? 'Critical Risk' : (reliableGap < 0 ? 'Planned PO Watch' : (issueLines.some(x => x.status === 'Critical Risk') ? 'Critical Risk' : (issueLines.some(x => x.status === 'Warehouse Prep Risk') ? 'Warehouse Prep Risk' : (issueLines.length ? 'Lead Time Watch' : 'OK'))));
    const action = totalGap < 0
      ? `Total demand is short by ${{Math.abs(totalGap)}} after all known POs; consider add-on PO or customer date change.`
      : (reliableGap < 0 ? `Current stock + confirmed incoming is short by ${{Math.abs(reliableGap)}}; coverage depends on planned PO or timing.` : 'Total demand is covered by current stock and confirmed incoming; watch line timing if any order is flagged.');
    rows.push({{
      status,
      sku,
      product:lines[0]?.product || base.product || '',
      onhand,
      total_demand:totalDemand,
      confirmed_incoming:confirmedIncoming,
      planned_incoming:plannedIncoming,
      reliable_gap:reliableGap,
      total_gap:totalGap,
      first_cover:firstCover,
      issue_orders:new Set(issueLines.map(x => x.so)).size,
      issue_lines:issueLines.length,
      action,
    }});
  }}
  return rows.sort((a,b) => statusRank(a.status)-statusRank(b.status) || Number(a.total_gap)-Number(b.total_gap) || String(a.sku).localeCompare(String(b.sku)));
}}
function buildSkuCoverageLadder(view) {{
  const overviewSkus = new Set([
    ...view.lines.map(x => x.sku),
    ...Object.keys(window.SCM_DATA.future_po_by_sku || {{}}),
  ]);
  const rows = [];
  for (const sku of overviewSkus) {{
    const lines = view.lines.filter(x => x.sku === sku);
    const base = skuInfoFor(sku);
    const totalDemand = lines.reduce((sum, x) => sum + Number(x.qty || 0), 0);
    let cumulative = Number(base.current_onhand || 0);
    rows.push({{status:cumulative >= totalDemand ? 'OK' : 'Lead Time Watch', sku, product:lines[0]?.product || base.product || '', step:'Current stock', eta:'', reliability:'Onhand', step_qty:cumulative, cumulative_supply:cumulative, total_demand:totalDemand, remaining_gap:cumulative-totalDemand, conclusion:cumulative >= totalDemand ? 'Current stock covers all demand.' : `Need ${{totalDemand-cumulative}} more units from incoming PO or add-on PO.`}});
    const pos = usableFuturePos(sku).map(x => ({{...x, qty:Number(x.qty || 0)}})).sort((a,b) => String(a.available || '9999-12-31').localeCompare(String(b.available || '9999-12-31')) || String(a.po).localeCompare(String(b.po)));
    for (const po of pos) {{
      cumulative += po.qty;
      const gap = cumulative - totalDemand;
      rows.push({{status:gap >= 0 ? 'OK' : (isPlannedReliability(po.reliability) ? 'Planned PO Watch' : 'Lead Time Watch'), sku, product:lines[0]?.product || base.product || '', step:po.po || '(blank PO)', eta:po.available || '', reliability:po.reliability || '', step_qty:po.qty, cumulative_supply:cumulative, total_demand:totalDemand, remaining_gap:gap, conclusion:gap >= 0 ? '累计供应已覆盖全部未完成 SO 需求。' : `该 PO 到仓后仍缺 ${{Math.abs(gap)}} 件。`}});
    }}
  }}
  return rows;
}}
function buildInboundNeedByRows(view) {{
  const rows = [];
  const bySku = new Map();
  for (const line of view.lines) {{
    if (!bySku.has(line.sku)) bySku.set(line.sku, []);
    bySku.get(line.sku).push(line);
  }}
  for (const [sku, lines] of bySku.entries()) {{
    const skuInfo = skuInfoFor(sku);
    const weeklyForecast = Number(skuInfo.weekly_forecast || 0);
    const dailyForecast = weeklyForecast / 7;
    const lots = usableFuturePos(sku).map((po, idx) => ({{
      lot_id:`${{po.po || '(blank PO)'}}__${{po.bol || ''}}__${{idx}}`,
      po:po.po || '',
      bol:po.bol || '',
      sku,
      product:skuInfo.product || '',
      qty:Number(po.qty || 0),
      remaining:Number(po.qty || 0),
      available:po.available || '',
      reliability:po.reliability || '',
      factory_date:po.factory_date || '',
      sailing_date:po.sailing_date || '',
      port_eta:po.port_eta || '',
      need_by:'',
      need_source:'Buffer Stock',
      first_affected_so:'',
      forecast_week:'',
      used_qty:0,
      reason:'Not needed within current SO and forecast horizon.',
    }})).sort((a,b) => String(a.available || '9999-12-31').localeCompare(String(b.available || '9999-12-31')) || String(a.po).localeCompare(String(b.po)));
    if (!lots.length) continue;
    let stock = Number(skuInfo.current_onhand || 0);
    const consumeFromLots = (qty, dateText, source, so='') => {{
      let need = Number(qty || 0);
      for (const lot of lots) {{
        if (need <= 0) break;
        if (lot.remaining <= 0) continue;
        const take = Math.min(lot.remaining, need);
        lot.remaining -= take;
        lot.used_qty += take;
        need -= take;
        if (!lot.need_by) {{
          lot.need_by = dateText;
          lot.need_source = source;
          lot.first_affected_so = so;
          lot.forecast_week = source === 'Forecast Needed' ? dateText : '';
          lot.reason = source === 'SO Needed' ? `First consumed by SO ${{so}}.` : 'First consumed by forecast after open SO demand.';
        }}
      }}
    }};
    const sortedSo = lines.slice().sort((a,b) => String(a.latest_ship || '9999-12-31').localeCompare(String(b.latest_ship || '9999-12-31')) || String(a.so).localeCompare(String(b.so)));
    const forecastConsumedByWeek = new Map();
    for (const line of sortedSo) {{
      let qty = Number(line.qty || 0);
      const stockUse = Math.min(Math.max(stock, 0), qty);
      stock -= stockUse;
      qty -= stockUse;
      if (line.latest_ship) {{
        const weekStart = addDays(line.latest_ship, -((new Date(`${{line.latest_ship}}T00:00:00`).getUTCDay() + 6) % 7));
        forecastConsumedByWeek.set(weekStart, (forecastConsumedByWeek.get(weekStart) || 0) + Number(line.qty || 0));
      }}
      const needByWh = line.latest_ship ? addDays(line.latest_ship, -7) : (line.required_arrival || window.SCM_DATA.generated_at);
      consumeFromLots(qty, needByWh, 'SO Needed', line.so);
    }}
    let forecastStock = stock;
    const horizonWeeks = 52;
    for (let w=1; w<=horizonWeeks; w++) {{
      const weekDate = addDays(window.SCM_DATA.generated_at, w * 7);
      const weekStart = addDays(weekDate, -((new Date(`${{weekDate}}T00:00:00`).getUTCDay() + 6) % 7));
      const soInWeek = forecastConsumedByWeek.get(weekStart) || 0;
      const forecastQty = Math.max(weeklyForecast - soInWeek, 0);
      if (forecastQty <= 0) continue;
      const stockUse = Math.min(Math.max(forecastStock, 0), forecastQty);
      forecastStock -= stockUse;
      consumeFromLots(forecastQty - stockUse, weekStart, 'Forecast Needed', '');
      if (lots.every(x => x.remaining <= 0 || x.need_by)) break;
    }}
    for (const lot of lots) {{
      const margin = lot.need_by && lot.available ? diffDays(lot.need_by, lot.available) : '';
      let status = 'Buffer Stock';
      let action = 'No urgent follow-up based on current SO and forecast.';
      if (lot.need_source === 'SO Needed' || lot.need_source === 'Forecast Needed') {{
        status = margin !== '' && margin < 0 ? 'Late Risk' : lot.need_source;
        action = margin !== '' && margin < 0 ? `ETA is late by ${{Math.abs(margin)}} days versus need-by date.` : `Need by ${{lot.need_by}} based on ${{lot.need_source === 'SO Needed' ? 'open SO' : 'forecast'}}.`;
      }}
      rows.push({{
        status,
        sku:lot.sku,
        product:lot.product,
        po:lot.po,
        bol:lot.bol,
        qty:Math.round(lot.qty),
        used_qty:Math.round(lot.used_qty),
        current_wh_eta:lot.available,
        need_by_wh_date:lot.need_by,
        need_source:lot.need_source,
        first_affected_so:lot.first_affected_so,
        forecast_week:lot.forecast_week,
        weekly_forecast:weeklyForecast,
        days_margin:margin,
        reliability:lot.reliability,
        factory_date:lot.factory_date,
        sailing_date:lot.sailing_date,
        port_eta:lot.port_eta,
        action,
      }});
    }}
  }}
  return rows.sort((a,b) => statusRank(a.status === 'Late Risk' ? 'Critical Risk' : a.status === 'SO Needed' ? 'Lead Time Watch' : a.status === 'Forecast Needed' ? 'Planned PO Watch' : 'OK') - statusRank(b.status === 'Late Risk' ? 'Critical Risk' : b.status === 'SO Needed' ? 'Lead Time Watch' : b.status === 'Forecast Needed' ? 'Planned PO Watch' : 'OK')
    || String(a.need_by_wh_date || '9999-12-31').localeCompare(String(b.need_by_wh_date || '9999-12-31'))
    || String(a.current_wh_eta || '9999-12-31').localeCompare(String(b.current_wh_eta || '9999-12-31'))
    || String(a.sku).localeCompare(String(b.sku)));
}}
function buildPortArrivalReminderRows() {{
  const groups = new Map();
  const shipmentByBol = new Map((window.SCM_DATA.shipment_control_rows || []).map(x => [String(x.bol || ''), x]));
  for (const po of (window.SCM_DATA.purchase_rows || [])) {{
    if (!po.bol || po.actual_receipt_date) continue;
    const shipment = shipmentByBol.get(String(po.bol)) || {{}};
    const portAtd = shipment.port_atd || po.shipment_port_atd || '';
    const portEta = shipment.port_eta || po.port_eta || '';
    const hasArrivedPort = !!portAtd || (!!portEta && portEta <= window.SCM_DATA.generated_at);
    if (!hasArrivedPort) continue;
    const key = String(po.bol);
    const row = groups.get(key) || {{
      status:'Critical Risk',
      bol:po.bol,
      related_pos:new Set(),
      sku_count:new Set(),
      qty:0,
      port_eta:portEta,
      port_atd:portAtd,
      planned_wh_eta:po.available_date,
      warehouse_ready_date:warehouseReadyDate(po.available_date),
      action:'已到港但 Purchase 还没有到仓 ATD，请催货代/仓库清关送仓；上传 Follow Up 后如果 Purchase ATD 有值，这条提醒会消失。',
    }};
    row.related_pos.add(po.po);
    row.sku_count.add(po.sku);
    row.qty += Number(po.qty || 0);
    if (!row.port_eta || (portEta && portEta < row.port_eta)) row.port_eta = portEta;
    if (!row.port_atd || (portAtd && portAtd < row.port_atd)) row.port_atd = portAtd;
    if (!row.planned_wh_eta || (po.available_date && po.available_date < row.planned_wh_eta)) row.planned_wh_eta = po.available_date;
    row.warehouse_ready_date = warehouseReadyDate(row.planned_wh_eta);
    groups.set(key, row);
  }}
  return [...groups.values()].map(row => ({{
    ...row,
    related_pos:[...row.related_pos].filter(Boolean).join('; '),
    sku_count:row.sku_count.size,
    qty:Math.round(row.qty),
  }})).sort((a,b) => String(a.port_atd || a.port_eta || '9999-12-31').localeCompare(String(b.port_atd || b.port_eta || '9999-12-31')) || String(a.bol).localeCompare(String(b.bol)));
}}
function buildWarehouseShipReminderRows(view) {{
  return view.soBoard
    .map(so => ({{
      ...so,
      days_to_ship: diffDays(so.latest_ship, window.SCM_DATA.generated_at),
      ship_notice_status: diffDays(so.latest_ship, window.SCM_DATA.generated_at) < 0 ? '已逾期' : (diffDays(so.latest_ship, window.SCM_DATA.generated_at) === 0 ? '今日需发货' : '10天内需发货'),
      action:'Sales 仍未填写 ATD，请确认美国仓是否已安排发货/配送；上传 Follow Up 后如果 Sales ATD 有值，这条提醒会消失。',
    }}))
    .filter(x => x.latest_ship && x.days_to_ship !== '' && x.days_to_ship <= 10)
    .sort((a,b) => Number(a.days_to_ship) - Number(b.days_to_ship) || statusRank(a.status) - statusRank(b.status) || String(a.so).localeCompare(String(b.so)));
}}
function buildAllInboundShipmentRows(view) {{
  const shipmentByBol = new Map((window.SCM_DATA.shipment_control_rows || []).map(x => [String(x.bol || ''), x]));
  const linesBySku = new Map();
  for (const line of view.lines) {{
    if (!linesBySku.has(line.sku)) linesBySku.set(line.sku, []);
    linesBySku.get(line.sku).push(line);
  }}
  const groups = new Map();
  for (const [sku, pos] of Object.entries(window.SCM_DATA.future_po_by_sku || {{}})) {{
    for (const po of usableFuturePos(sku)) {{
      if (!po.bol) continue;
      const shipment = shipmentByBol.get(String(po.bol)) || {{}};
      const portEta = shipment.port_eta || po.port_eta || '';
      const plannedWhEta = po.available || (portEta ? addDays(portEta, 7) : '');
      const g = groups.get(po.bol) || {{
        status:'OK',
        bol:po.bol,
        related_pos:new Set(),
        sku_count:new Set(),
        qty:0,
        sailing_date:shipment.sailing_date || po.sailing_date || '',
        port_eta:portEta,
        port_atd:shipment.port_atd || po.shipment_port_atd || '',
        planned_wh_eta:plannedWhEta,
        warehouse_ready_date:warehouseReadyDate(plannedWhEta),
        earliest_so:'',
        earliest_latest_ship:'',
        days_margin:'',
        action:'按当前 SO 时间暂无紧急风险，继续跟进船期和送仓。',
      }};
      g.related_pos.add(po.po || '');
      g.sku_count.add(sku);
      g.qty += Number(po.qty || 0);
      if (!g.sailing_date || ((shipment.sailing_date || po.sailing_date) && (shipment.sailing_date || po.sailing_date) < g.sailing_date)) g.sailing_date = shipment.sailing_date || po.sailing_date || g.sailing_date;
      if (!g.port_eta || (portEta && portEta < g.port_eta)) g.port_eta = portEta;
      if (!g.port_atd || ((shipment.port_atd || po.shipment_port_atd) && (shipment.port_atd || po.shipment_port_atd) < g.port_atd)) g.port_atd = shipment.port_atd || po.shipment_port_atd || g.port_atd;
      if (!g.planned_wh_eta || (plannedWhEta && plannedWhEta < g.planned_wh_eta)) g.planned_wh_eta = plannedWhEta;
      g.warehouse_ready_date = warehouseReadyDate(g.planned_wh_eta);
      for (const line of (linesBySku.get(sku) || [])) {{
        if (!g.earliest_latest_ship || (line.latest_ship && line.latest_ship < g.earliest_latest_ship)) {{
          g.earliest_latest_ship = line.latest_ship;
          g.earliest_so = line.so;
        }}
      }}
      groups.set(po.bol, g);
    }}
  }}
  return [...groups.values()].map(g => {{
    g.days_margin = g.earliest_latest_ship && g.warehouse_ready_date ? diffDays(g.earliest_latest_ship, g.warehouse_ready_date) : '';
    if (g.days_margin !== '' && g.days_margin < 0) {{
      g.status = 'Critical Risk';
      g.action = `仓库7天备货后会晚 ${{Math.abs(g.days_margin)}} 天，需催送仓/卸货或沟通客户交期。`;
    }} else if (g.days_margin !== '' && g.days_margin <= 3) {{
      g.status = 'Lead Time Watch';
      g.action = '仓库备货余量很少，请跟进清关、送仓和卸货节奏。';
    }}
    return {{
      ...g,
      related_pos:[...g.related_pos].filter(Boolean).join('; '),
      sku_count:g.sku_count.size,
      qty:Math.round(g.qty),
    }};
  }}).sort((a,b) => statusRank(a.status)-statusRank(b.status) || String(a.port_eta || '9999-12-31').localeCompare(String(b.port_eta || '9999-12-31')) || String(a.bol).localeCompare(String(b.bol)));
}}
function renderDynamicViews() {{
  const view = currentViewData();
  const manualAlerts = view.audit.filter(x => x.status === 'Unallocated Stock');
  const emailRows = warehouseEmailRows(view);
  const logisticsRows = logisticsActionRows(view);
  const skuRows = buildSkuShortageRows(view);
  const skuCoverageRows = buildSkuCoverageOverview(view);
  const skuCoverageLadder = buildSkuCoverageLadder(view);
  const needByRows = buildInboundNeedByRows(view);
  const allInboundRows = buildAllInboundShipmentRows(view);
  const portArrivalRows = buildPortArrivalReminderRows();
  const whShipRows = buildWarehouseShipReminderRows(view);
  const dueEmailRows = emailRows.filter(x => ['Overdue','Due Today','Upcoming'].includes(x.notice_status));
  document.getElementById('cardOpenSo').textContent = view.soBoard.length;
  document.getElementById('cardCriticalSo').textContent = view.soBoard.filter(x => x.status === 'Critical Risk').length;
  document.getElementById('cardWatchSo').textContent = view.soBoard.filter(x => ['Warehouse Prep Risk','Lead Time Watch','Planned PO Watch'].includes(x.status)).length;
  document.getElementById('cardIssueSku').textContent = skuRows.filter(x => x.status !== 'OK' || x.shortage_qty > 0 || x.issue_qty > 0).length;
  document.getElementById('cardActionQueue').textContent = view.actionQueue.length + portArrivalRows.length + whShipRows.length + dueEmailRows.length;
  document.getElementById('cardDataChecks').textContent = window.SCM_DATA.summary?.data_check_count ?? window.SCM_DATA.summary?.old_open_line_count ?? 0;
  const notice = manualAlerts.length ? simplePanel('人工分配检查', manualAlerts, [['status','状态'],['sku','SKU'],['product','产品'],['current_onhand','当前库存'],['manual_allocated','人工分配数量'],['remaining_stock','剩余库存'],['remaining_risk_qty','剩余风险数量'],['possible_targets','可调整 SO'],['action','处理建议']], '这些 SKU 仍有剩余库存，可能可以继续降低订单风险。') : '';
  const emailNotice = dueEmailRows.length ? warehouseEmailPanel(dueEmailRows, '仓库邮件到期确认', '发送提前 21 天仓库通知后，在这里标记已发送。') : '';
  const soRiskToday = simplePanel('今日 SO 风险提醒', view.actionQueue, [['status','状态'],['so','SO/CI'],['customer','客户'],['delivery_center','客户仓'],['required_arrival','客户要求到仓日'],['latest_ship','最晚美仓发货日'],['suggested_customer_date','建议客户交期'],['line_count','总SKU数'],['critical_skus','高风险SKU数'],['warehouse_skus','仓库备货风险SKU数'],['watch_skus','确认在途SKU数'],['planned_skus','预计在途SKU数'],['action','处理建议']], '只显示今天需要关注的订单：高风险 SO，以及 21 天内需要处理的观察 SO。');
  const portArrivalNotice = simplePanel('在途船期跟进提醒', portArrivalRows, [['status','状态'],['bol','BOL'],['related_pos','相关PO'],['sku_count','SKU数'],['qty','数量'],['port_eta','预计到港日'],['port_atd','实际到港日'],['planned_wh_eta','预计到仓日'],['warehouse_ready_date','仓库备货后可发货日'],['action','处理建议']], '每船一条。Shipment Control 已显示到港，但 Purchase 还没有到仓 ATD，需要催清关/送仓/卸货。');
  const whShipNotice = simplePanel('SO 美仓发货提醒（10天内）', whShipRows, [['ship_notice_status','提醒状态'],['so','SO/CI'],['customer','客户'],['delivery_center','客户仓'],['required_arrival','客户要求到仓日'],['latest_ship','最晚美仓发货日'],['days_to_ship','剩余天数'],['status','风险'],['issue_skus','问题SKU数'],['action','处理建议']], 'Sales 里 ATD 为空且最晚美仓发货日在 10 天内；上传 Follow Up 后如果 Sales ATD 有值，提醒会自动取消。');
  document.getElementById('today').innerHTML = emailNotice + soRiskToday + portArrivalNotice + whShipNotice;
  if (notice) document.getElementById('today').innerHTML = notice + document.getElementById('today').innerHTML;
  document.getElementById('so').innerHTML =
    simplePanel('订单层级看板', view.soBoard, [['status','状态'],['so','SO/CI'],['customer','客户'],['delivery_center','客户仓'],['latest_ship','最晚美仓发货日'],['required_arrival','客户要求到仓日'],['suggested_customer_date','建议客户交期'],['line_count','行数'],['issue_skus','问题SKU数'],['critical_skus','高风险SKU数'],['warehouse_skus','仓库备货风险SKU数'],['planned_skus','预计在途SKU数'],['watch_skus','确认在途观察SKU数'],['action','处理建议']], '已考虑人工分配。');
  document.getElementById('control').innerHTML = simplePanel('订单 SKU 明细', view.lines.sort((a,b)=>statusRank(a.status)-statusRank(b.status)||String(a.latest_ship).localeCompare(String(b.latest_ship))||String(a.so).localeCompare(String(b.so))), [['status','状态'],['so','SO/CI'],['customer','客户'],['delivery_center','客户仓'],['sku','SKU'],['product','产品'],['qty','数量'],['risk_qty','人工风险数量'],['system_risk_qty','系统风险数量'],['manual_stock_qty','人工分配库存'],['allocation_issue_skus','SO问题SKU数'],['allocation_critical_skus','SO高风险SKU数'],['latest_ship','最晚美仓发货日'],['required_arrival','客户要求到仓日'],['suggested_customer_date','建议客户交期'],['cover_type','覆盖类型'],['cover','覆盖来源'],['action','处理建议']], '已考虑人工分配。');
  const auditPanel = simplePanel('人工分配检查', view.audit, [['status','状态'],['sku','SKU'],['product','产品'],['current_onhand','当前库存'],['manual_allocated','人工分配数量'],['remaining_stock','剩余库存'],['remaining_risk_qty','剩余风险数量'],['decision_rows','决策行数'],['possible_targets','可调整 SO'],['action','处理建议']], '显示人工分配后是否仍有剩余库存或风险。');
  document.getElementById('sku').innerHTML =
    simplePanel('SKU 缺货跟进', skuRows, [['status','状态'],['sku','SKU'],['product','产品'],['current_onhand','当前库存'],['open_demand','未交货需求'],['confirmed_incoming_qty','确认在途'],['planned_incoming_qty','预计在途'],['shortage_qty','缺口数量'],['issue_qty','风险数量'],['follow_po','跟进PO'],['follow_bol','跟进BOL'],['port_eta','ETD到港日'],['cover_eta','预计到仓日'],['warehouse_ready_date','仓库备货后可发货日'],['latest_wh_arrival','最晚美仓发货日'],['gap_days','差异天数'],['affected_so_count','影响SO数'],['affected_sos','影响SO'],['action','处理建议']], '库存、需求、在途、缺口，以及需要跟进的 PO/BOL。差异天数优先按“仓库备货后可发货日”计算。') +
    simplePanel('SKU 覆盖阶梯', skuCoverageLadder, [['status','状态'],['sku','SKU'],['product','产品'],['step','库存/PO'],['eta','预计到仓'],['reliability','可靠性'],['step_qty','数量'],['cumulative_supply','累计供应'],['total_demand','未交货需求'],['remaining_gap','该节点后缺口'],['conclusion','结论']], '按 SKU 从上往下看：现有库存和每个 PO 到仓后，何时能完全覆盖未交货需求。') +
    auditPanel;
  const factoryCols = [['po','PO'],['urgent_skus','紧急SKU'],['urgent_sos','紧急SO'],['earliest_required_arrival','最早客户交期'],['latest_wh_ship','最晚美仓发货日'],['latest_wh_arrival','最晚到仓红线'],['latest_factory_ship','最晚工厂发货日'],['planned_factory_ship','计划工厂发货日'],['current_wh_eta','当前预计到仓日'],['gap_days','差异天数'],['affected_so_count','影响SO数'],['affected_sku_count','影响SKU数'],['qty','数量'],['action','处理建议']];
  const inboundCols = [['status','状态'],['bol','BOL'],['related_pos','相关PO'],['sku_count','SKU数'],['qty','数量'],['sailing_date','船期'],['port_eta','ETD到港日'],['port_atd','实际到港日'],['planned_wh_eta','预计到仓日'],['warehouse_ready_date','仓库备货后可发货日'],['earliest_so','最早影响SO'],['earliest_latest_ship','最早SO最晚发货日'],['days_margin','备货后余量'],['action','处理建议']];
  const bolCols = [['bol','BOL'],['related_pos','相关PO'],['urgent_line_count','紧急行数'],['later_so_count','后续SO数'],['urgent_skus','紧急SKU'],['earliest_required_arrival','最早客户交期'],['latest_wh_ship','最晚美仓发货日'],['port_eta','ETD到港日'],['planned_wh_eta','计划到仓日'],['warehouse_ready_date','仓库备货后可发货日'],['warehouse_days_left','仓库备货后余量'],['suggested_customer_date','建议客户交期'],['affected_so_count','影响SO数'],['affected_sku_count','影响SKU数'],['qty','数量'],['forwarder_action','货代处理建议'],['warehouse_action','仓库处理建议']];
  const bolLineCols = [['bol','BOL'],['po','PO'],['sku','SKU'],['so','SO/CI'],['customer','客户'],['delivery_center','客户仓'],['qty','数量'],['required_arrival','客户要求到仓日'],['latest_wh_ship','最晚美仓发货日'],['port_eta','ETD到港日'],['planned_wh_eta','计划到仓日'],['warehouse_ready_date','仓库备货后可发货日'],['warehouse_days_left','仓库备货后余量'],['status','状态'],['action','处理建议']];
  document.getElementById('logistics').innerHTML =
    simplePanel('所有在途船期总览', allInboundRows, inboundCols, '每个 BOL 一行：显示 ETD 到港、预计到仓、仓库 7 天备货后可发货日，以及最早受影响 SO 的最晚发货日。') +
    simplePanel('工厂发货船期催办', logisticsRows.filter(x => x.category === 'Factory PO Follow-up'), factoryCols, '按 PO 跟进供应商：哪些 SKU/SO 受影响，以及最晚需要工厂什么时候发货。') +
    simplePanel('提单清关送仓汇总', logisticsRows.filter(x => x.category === 'BOL Inbound Summary'), bolCols, '一行一个 BOL。点击 BOL 可筛选下方紧急明细。', 'bolSummary') +
    '<div class="inline-filter" id="bolInlineFilter"><span id="bolInlineFilterText"></span><button id="clearBolInlineFilter">关闭 BOL 筛选</button></div>' +
    simplePanel('提单紧急明细', logisticsRows.filter(x => x.category === 'BOL-SKU-SO Urgent Detail'), bolLineCols, '显示行动窗口内每条紧急 BOL-SKU-SO；较远期 SO 计入后续 SO 数。', 'bolUrgentDetail');
  document.getElementById('needby').innerHTML = simplePanel('在途最晚到仓红线', needByRows, [['status','状态'],['sku','SKU'],['product','产品'],['po','PO'],['bol','BOL'],['qty','数量'],['used_qty','已占用数量'],['current_wh_eta','当前预计到仓日'],['need_by_wh_date','最晚需要到仓日'],['need_source','需求来源'],['first_affected_so','最早影响SO'],['forecast_week','预测周'],['weekly_forecast','周预测'],['days_margin','ETA余量'],['reliability','可靠性'],['factory_date','工厂日期'],['sailing_date','船期'],['port_eta','ETD到港日'],['action','处理建议']], 'SO 需求的最晚需要到仓日 = SO 最晚美仓发货日 - 7 天仓库备货。ETA余量 = 最晚需要到仓日 - 当前预计到仓日；负数表示晚到，需要跟进。');
  renderTransitSettings();
  renderImportNotice();
}}
function simplePanel(title, items, cols, subtitle='', id='') {{
  return `<section class="panel"${{id ? ` id="${{esc(id)}}"` : ''}}><div class="panel-title"><h2>${{esc(title)}}</h2><p>${{esc(subtitle)}}</p></div>${{simpleTable(items, cols)}}</section>`;
}}
function logisticsActionRows(view) {{
  const poGroups = new Map();
  const bolGroups = new Map();
  const bolLineRows = [];
  for (const line of view.lines) {{
    if (line.status === 'OK') continue;
    for (const cover of (line.cover_details || [])) {{
      const hasConfirmedShipment = !!(cover.bol || cover.sailing_date);
      if (!hasConfirmedShipment) {{
        const poKey = cover.po || '(blank PO)';
        const pg = poGroups.get(poKey) || {{
          category:'Factory PO Follow-up',
          po:poKey,
          urgent_skus:new Set(),
          urgent_sos:new Set(),
          affected_sos:new Set(),
          affected_skus:new Set(),
          earliest_required_arrival:'',
          latest_wh_ship:'',
          latest_wh_arrival:'',
          latest_factory_ship:'',
          planned_factory_ship:'',
          current_wh_eta:'',
          gap_days:'',
          qty:0,
          action:'',
        }};
        pg.urgent_skus.add(line.sku);
        pg.urgent_sos.add(line.so);
        pg.affected_sos.add(line.so);
        pg.affected_skus.add(line.sku);
        pg.qty += Number(cover.qty || 0);
        if (!pg.earliest_required_arrival || (line.required_arrival && line.required_arrival < pg.earliest_required_arrival)) pg.earliest_required_arrival = line.required_arrival;
        if (!pg.latest_wh_ship || (line.latest_ship && line.latest_ship < pg.latest_wh_ship)) pg.latest_wh_ship = line.latest_ship;
        const factoryNeedWhArrival = line.latest_ship ? addDays(line.latest_ship, -7) : '';
        if (!pg.latest_wh_arrival || (factoryNeedWhArrival && factoryNeedWhArrival < pg.latest_wh_arrival)) pg.latest_wh_arrival = factoryNeedWhArrival;
        if (!pg.current_wh_eta || (cover.available && cover.available < pg.current_wh_eta)) pg.current_wh_eta = cover.available;
        if (!pg.planned_factory_ship || (cover.factory_date && cover.factory_date < pg.planned_factory_ship)) pg.planned_factory_ship = cover.factory_date;
        poGroups.set(poKey, pg);
      }}

      if (cover.bol) {{
        const bg = bolGroups.get(cover.bol) || {{
          category:'BOL Inbound Summary',
          bol:cover.bol,
          related_pos:new Set(),
          urgent_skus:new Set(),
          affected_sos:new Set(),
          affected_skus:new Set(),
          later_sos:new Set(),
          earliest_required_arrival:'',
          latest_wh_ship:'',
          port_eta:'',
          planned_wh_eta:'',
          warehouse_ready_date:'',
          warehouse_days_left:'',
          suggested_customer_date:'',
          urgent_line_count:0,
          later_so_count:0,
          qty:0,
          forwarder_action:'',
          warehouse_action:'',
        }};
        const whReady = warehouseReadyDate(cover.available || '');
        const whDaysLeft = line.latest_ship && whReady ? diffDays(line.latest_ship, whReady) : '';
        const daysToShip = diffDays(line.latest_ship, window.SCM_DATA.generated_at);
        const isUrgentLine = line.status === 'Critical Risk' || (whDaysLeft !== '' && whDaysLeft <= 7) || (daysToShip !== '' && daysToShip <= 21);
        bg.related_pos.add(cover.po || '');
        bg.affected_sos.add(line.so);
        bg.affected_skus.add(line.sku);
        bg.qty += Number(cover.qty || 0);
        if (isUrgentLine) {{
          bg.urgent_skus.add(line.sku);
          bg.urgent_line_count += 1;
          if (!bg.earliest_required_arrival || (line.required_arrival && line.required_arrival < bg.earliest_required_arrival)) bg.earliest_required_arrival = line.required_arrival;
          if (!bg.latest_wh_ship || (line.latest_ship && line.latest_ship < bg.latest_wh_ship)) bg.latest_wh_ship = line.latest_ship;
          bolLineRows.push({{
            category:'BOL-SKU-SO Urgent Detail',
            bol:cover.bol,
            po:cover.po || '',
            sku:line.sku,
            so:line.so,
            customer:line.customer,
            delivery_center:line.delivery_center,
            qty:Math.round(Number(cover.qty || line.qty || 0)),
            required_arrival:line.required_arrival,
            latest_wh_ship:line.latest_ship,
            port_eta:cover.port_eta || '',
            planned_wh_eta:cover.available || '',
            warehouse_ready_date:whReady,
            warehouse_days_left:whDaysLeft,
            status:line.status,
            action: whDaysLeft !== '' && whDaysLeft < 0 ? `Late by ${{Math.abs(whDaysLeft)}} days; communicate customer date or expedite receiving.` : (whDaysLeft !== '' && whDaysLeft <= 2 ? 'Very tight: prioritize clearance, delivery, receiving, and release.' : 'Monitor BOL receiving and release before latest ship date.'),
          }});
        }} else {{
          bg.later_sos.add(line.so);
        }}
        if (!bg.port_eta || (cover.port_eta && cover.port_eta < bg.port_eta)) bg.port_eta = cover.port_eta;
        if (!bg.planned_wh_eta || (cover.available && cover.available < bg.planned_wh_eta)) bg.planned_wh_eta = cover.available;
        bg.warehouse_ready_date = warehouseReadyDate(bg.planned_wh_eta);
        if (line.suggested_customer_date && (!bg.suggested_customer_date || line.suggested_customer_date > bg.suggested_customer_date)) bg.suggested_customer_date = line.suggested_customer_date;
        bolGroups.set(cover.bol, bg);
      }}
    }}
  }}
  const factoryRows = [...poGroups.values()].map(g => {{
    g.latest_factory_ship = latestFactoryDate(g.latest_wh_arrival);
    g.gap_days = g.current_wh_eta && g.latest_wh_arrival ? diffDays(g.current_wh_eta, g.latest_wh_arrival) : '';
    g.action = g.gap_days !== '' && g.gap_days > 0
      ? `Need PO/factory timing earlier by ${{g.gap_days}} days, or communicate customer date.`
      : 'Confirm factory shipment, sailing date, and BOL.';
    return {{
      ...g,
      urgent_skus:[...g.urgent_skus].slice(0,12).join('; '),
      urgent_sos:[...g.urgent_sos].slice(0,12).join('; '),
      affected_so_count:g.affected_sos.size,
      affected_sku_count:g.affected_skus.size,
      qty:Math.round(g.qty),
    }};
  }});
  const bolRows = [...bolGroups.values()].map(g => {{
    g.warehouse_ready_date = warehouseReadyDate(g.planned_wh_eta);
    g.warehouse_days_left = g.latest_wh_ship && g.warehouse_ready_date ? diffDays(g.latest_wh_ship, g.warehouse_ready_date) : '';
    g.later_so_count = g.later_sos.size;
    g.forwarder_action = g.warehouse_days_left !== '' && g.warehouse_days_left < 3
      ? 'Urgent: push customs clearance, truck appointment, and delivery to WH.'
      : 'Confirm clearance, appointment, and delivery date.';
    g.warehouse_action = g.warehouse_days_left !== '' && g.warehouse_days_left < 3
      ? 'Ask warehouse to prioritize receiving/release immediately after delivery.'
      : 'Ask warehouse to plan receiving before latest ship date.';
    return {{
      ...g,
      related_pos:[...g.related_pos].filter(Boolean).join('; '),
      urgent_skus:[...g.urgent_skus].slice(0,12).join('; '),
      affected_so_count:g.affected_sos.size,
      affected_sku_count:g.affected_skus.size,
      qty:Math.round(g.qty),
    }};
  }});
  return [...factoryRows, ...bolRows, ...bolLineRows].sort((a,b) => {{
    const catRank = {{'Factory PO Follow-up':0,'BOL Inbound Summary':1,'BOL-SKU-SO Urgent Detail':2}};
    return (catRank[a.category] ?? 9) - (catRank[b.category] ?? 9)
      || String(a.latest_wh_ship || '9999-12-31').localeCompare(String(b.latest_wh_ship || '9999-12-31'))
      || String(a.bol || a.po).localeCompare(String(b.bol || b.po))
      || String(a.sku || '').localeCompare(String(b.sku || ''))
      || String(a.so || '').localeCompare(String(b.so || ''));
  }});
}}
function showDetail(kind, value) {{
  const d = window.SCM_DATA;
  const detail = document.getElementById('detail');
  let html = '';
  if (kind === 'so') {{
    const view = currentViewData();
    const so = view.soBoard.find(x => x.so === value);
    const lines = view.lines.filter(x => x.so === value);
    if (!so) return;
    html = `<div class="detail-head"><div><h2>SO ${{esc(value)}} · ${{statusBadge(so.status)}}</h2><p>${{esc(so.customer)}} · ${{esc(so.delivery_center)}} · 最晚美仓发货日 ${{esc(so.latest_ship)}} · 客户要求到仓日 ${{esc(so.required_arrival)}}</p></div><button id="closeDetail">关闭</button></div>
      <div class="detail-grid">${{mini('客户要求到仓日', so.required_arrival)}}${{mini('最晚美仓发货日', so.latest_ship)}}${{mini('行数', so.line_count)}}${{mini('问题SKU数', so.issue_skus)}}${{mini('高风险SKU数', so.critical_skus)}}${{mini('预计在途SKU数', so.planned_skus)}}${{mini('确认在途观察SKU数', so.watch_skus)}}${{mini('建议客户交期', so.suggested_customer_date || '')}}</div>
      <div class="detail-body">${{simpleTable(lines, [['status','状态'],['sku','SKU'],['product','产品'],['qty','数量'],['risk_qty','人工风险数量'],['system_risk_qty','系统风险数量'],['manual_stock_qty','人工分配库存'],['allocation_issue_skus','SO问题SKU数'],['allocation_critical_skus','SO高风险SKU数'],['latest_ship','最晚美仓发货日'],['required_arrival','客户要求到仓日'],['suggested_customer_date','建议客户交期'],['cover_type','覆盖类型'],['cover','覆盖来源'],['action','处理建议']])}}</div>`;
  }}
  if (kind === 'sku') {{
    const sku = skuInfoFor(value);
    const view = currentViewData();
    const lines = view.lines.filter(x => x.sku === value);
    const rec = d.recommendations.filter(x => x.sku === value);
    if (!sku) return;
    const confirmedRows = confirmedSpsAsSalesLines().filter(x => x.sku === value);
    const spsRows = lastSpsRiskPreview.filter(x => x.sku === value);
    const spsDemand = confirmedRows.length ? confirmedRows.reduce((sum, x) => sum + Number(x.qty || 0), 0) : spsRows.reduce((sum, x) => sum + Number(x.qty || 0), 0);
    const spsExplainRows = confirmedRows.length ? confirmedRows.map(x => ({{status:'Confirmed', order:x.so, qty:x.qty, etd:x.required_arrival, required_arrival:x.required_arrival, latest_ship:x.latest_ship, cover:'Included in local dashboard', uncovered:'', why:`当前库存 ${{sku.current_onhand}} - 已有需求 ${{sku.open_demand}} - 已确认 SPS 需求 ${{spsDemand}} = ${{Number(sku.current_onhand || 0) - Number(sku.open_demand || 0) - spsDemand}}`}})) : spsRows;
    const spsExplain = spsExplainRows.length ? `<section class="panel"><div class="panel-title"><h2>SPS 新单影响</h2><p>该 SKU 已确认/已导入的 SPS 行。</p></div><div class="detail-grid">${{mini('当前库存', sku.current_onhand)}}${{mini('已有需求', sku.open_demand)}}${{mini('SPS需求', spsDemand)}}${{mini('导入后缺口', Number(sku.current_onhand || 0) - Number(sku.open_demand || 0) - spsDemand)}}</div>${{simpleTable(spsExplainRows, [['status','状态'],['order','SO/PO'],['qty','数量'],['etd','ETD'],['latest_ship','最晚美仓发货日'],['required_arrival','客户要求到仓日'],['cover','覆盖来源'],['uncovered','未覆盖数量'],['why','说明']])}}</section>` : '';
    html = `<div class="detail-head"><div><h2>SKU ${{esc(value)}}</h2><p>${{esc(sku.product)}}</p></div><button id="closeDetail">关闭</button></div>
      <div class="detail-grid">${{mini('当前库存', sku.current_onhand)}}${{mini('未交货需求', sku.open_demand)}}${{mini('库存缺口', sku.stock_gap)}}${{mini('风险数量', sku.issue_qty)}}${{mini('高风险数量', sku.critical_qty)}}${{mini('观察数量', sku.watch_qty)}}${{mini('未来PO数量', sku.future_qty)}}${{mini('未覆盖数量', sku.uncovered_qty)}}</div>
      <div class="detail-body">${{rec.length ? simpleTable(rec, [['sku','SKU'],['stock_gap','库存缺口'],['issue_qty','风险数量'],['standard_affected_so','自然影响SO'],['recommended_so','建议承担SO'],['recommendation','建议'],['recommended_reason','原因']]) : ''}}
      ${{spsExplain}}
      ${{allocationTable(value, lines)}}
      ${{simpleTable(lines, [['status','状态'],['so','SO/CI'],['customer','客户'],['delivery_center','客户仓'],['qty','数量'],['risk_qty','风险数量'],['allocation_issue_skus','SO问题SKU数'],['allocation_critical_skus','SO高风险SKU数'],['latest_ship','最晚美仓发货日'],['required_arrival','客户要求到仓日'],['cover','覆盖来源'],['action','处理建议']])}}</div>`;
  }}
  if (kind === 'po') {{
    const po = d.po_impact.find(x => x.po === value);
    if (!po) return;
    const lines = po.priority_lines || [];
    html = `<div class="detail-head"><div><h2>PO ${{esc(value)}}</h2><p>根据关联 SO 的最晚美仓发货日，判断卸货/入库优先级。</p></div><button id="closeDetail">关闭</button></div>
      <div class="detail-grid">${{mini('覆盖SKU数', po.sku_count)}}${{mini('覆盖SO数', po.so_count)}}${{mini('高风险行数', po.critical_lines)}}${{mini('预计在途行数', po.planned_lines || 0)}}${{mini('确认在途观察行数', po.watch_lines)}}${{mini('延误数量', po.late_qty)}}</div>
      <div class="detail-body">${{simpleTable(lines, [['status','状态'],['reliability','PO可靠性'],['sku','SKU'],['product','产品'],['so','SO/CI'],['qty','数量'],['available','预计到仓'],['latest_ship','最晚美仓发货日'],['required_arrival','客户要求到仓日'],['days_delta','延误天数'],['delivery_center','客户仓']])}}</div>`;
  }}
  detail.innerHTML = html;
  detail.classList.add('active');
  document.getElementById('closeDetail')?.addEventListener('click', () => detail.classList.remove('active'));
  detail.scrollIntoView({{behavior:'smooth', block:'start'}});
}}
const buttons = [...document.querySelectorAll('button[data-tab]')];
const tabs = [...document.querySelectorAll('.tab')];
buttons.forEach(btn => btn.addEventListener('click', () => {{
  buttons.forEach(b => b.classList.toggle('active', b === btn));
  tabs.forEach(t => t.classList.toggle('active', t.id === btn.dataset.tab));
}}));
function filterAndFocusBol(value) {{
  document.querySelector('button[data-tab="logistics"]')?.click();
  activeFilter = value ? {{kind:'BOL', value}} : null;
  const search = document.getElementById('search');
  if (search) {{
    search.value = value || '';
  }}
  applyGlobalFilter();
  setTimeout(() => document.getElementById('bolUrgentDetail')?.scrollIntoView({{behavior:'smooth', block:'start'}}), 0);
}}
let activeFilter = null;
function updateFilterBanner() {{
  const banner = document.getElementById('activeFilterBanner');
  const text = document.getElementById('activeFilterText');
  const inline = document.getElementById('bolInlineFilter');
  const inlineText = document.getElementById('bolInlineFilterText');
  if (!banner || !text) return;
  if (!activeFilter) {{
    banner.style.display = 'none';
    text.innerHTML = '';
    if (inline) inline.style.display = 'none';
    if (inlineText) inlineText.innerHTML = '';
    return;
  }}
  banner.style.display = 'flex';
  text.innerHTML = `Viewing <b>${{esc(activeFilter.kind)}} ${{esc(activeFilter.value)}}</b>. Clear filter to return to all rows.`;
  if (inline) inline.style.display = 'flex';
  if (inlineText) inlineText.innerHTML = `Viewing <b>${{esc(activeFilter.kind)}} ${{esc(activeFilter.value)}}</b>.`;
}}
function applyGlobalFilter() {{
  const search = document.getElementById('search');
  const q = String(search?.value || '').trim().toLowerCase();
  document.querySelectorAll('tbody tr').forEach(tr => {{
    tr.style.display = !q || tr.innerText.toLowerCase().includes(q) ? '' : 'none';
  }});
  updateFilterBanner();
}}
function csvEscape(v) {{
  if (Array.isArray(v)) v = v.map(x => typeof x === 'object' ? (x.sku || x.so || x.po || '') : x).filter(Boolean).join('; ');
  else if (v && typeof v === 'object') v = v.sku || v.so || v.po || v.status || '';
  const s = String(v ?? '');
  const needsQuote = s.includes('"') || s.includes(',') || s.includes(String.fromCharCode(10)) || s.includes(String.fromCharCode(13));
  return needsQuote ? `"${{s.replace(/"/g, '""')}}"` : s;
}}
const EXPORT_COLS = {{
  action_queue:['reason','status','so','customer','delivery_center','latest_ship','days_to_ship','required_arrival','suggested_customer_date','line_count','issue_skus','critical_skus','planned_skus','watch_skus','confirmed_po_skus','action'],
  so_board:['status','so','customer','delivery_center','latest_ship','required_arrival','suggested_customer_date','line_count','issue_skus','critical_skus','planned_skus','watch_skus','confirmed_po_skus','action'],
  so_lines:['status','system_status','so','customer','delivery_center','sku','product','qty','risk_qty','system_risk_qty','manual_stock_qty','allocation_issue_skus','allocation_critical_skus','latest_ship','required_arrival','suggested_customer_date','stock_before','stock_used','cover_type','cover','uncovered','action'],
  sku_summary:['status','sku','product','current_onhand','open_demand','confirmed_incoming_qty','planned_incoming_qty','shortage_qty','issue_qty','follow_po','follow_bol','cover_eta','latest_wh_arrival','gap_days','affected_so_count','affected_sos','action'],
  replenishment:['status','sku','product','weekly_forecast','lead_time_weeks','safety_stock_weeks','buffer_weeks','target_weeks','target_stock','total_stock','formula_gap','case_pack','suggested_po_qty','sum_purchase_suggestion','order_risk_gap','onhand','open_demand','confirmed_incoming','planned_incoming','moq','action'],
  po_impact:['po','sku_count','so_count','critical_lines','planned_lines','watch_lines','late_qty'],
  logistics_actions:['category','po','bol','related_pos','sku','so','customer','delivery_center','urgent_line_count','later_so_count','urgent_skus','earliest_required_arrival','required_arrival','latest_wh_ship','latest_wh_arrival','latest_factory_ship','planned_factory_ship','current_wh_eta','port_eta','planned_wh_eta','warehouse_days_left','suggested_customer_date','affected_so_count','affected_sku_count','qty','gap_days','status','forwarder_action','warehouse_action','action'],
  inbound_need_by:['status','sku','product','po','bol','qty','used_qty','current_wh_eta','need_by_wh_date','need_source','first_affected_so','forecast_week','weekly_forecast','days_margin','reliability','factory_date','sailing_date','port_eta','action'],
  manual_audit:['status','sku','product','current_onhand','manual_allocated','remaining_stock','remaining_risk_qty','decision_rows','possible_targets','action'],
  warehouse_email:['notice_status','so','customer','delivery_center','required_arrival','email_due_date','days_to_email','latest_ship','risk_status','line_count','issue_skus','sent_at','action'],
  confirmed_sps_sales_rows:['Date','Item','Product Code','SO/CI','Zoho Order','ORD QTY','Unit Price','Total Sales','Customer','Delivery Center','ETD','ATD','Product Cost','Total Cost','Sales Order Update'],
}};
function projectRows(items, cols) {{
  if (!cols) return items;
  return items.map(row => Object.fromEntries(cols.map(col => [col, row[col] ?? ''])));
}}
function toCsv(items, cols=null) {{
  if (!items.length) return '';
  const csvCols = cols || [...new Set(items.flatMap(x => Object.keys(x)).filter(k => !['skus','raw','cover_pos','priority_lines'].includes(k)))];
  const rows = projectRows(items, csvCols);
  return [csvCols.join(','), ...rows.map(row => csvCols.map(c => csvEscape(row[c])).join(','))].join(String.fromCharCode(10));
}}
function downloadCsv(name, items, cols=null) {{
  const csv = toCsv(items, cols);
  const blob = new Blob(['\ufeff' + csv], {{type:'text/csv;charset=utf-8'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  const stamp = new Date().toISOString().slice(0,19).replace(/[-:T]/g,'');
  a.download = name.replace(/\\.csv$/i, `_clean_${{stamp}}.csv`);
  a.click();
  URL.revokeObjectURL(a.href);
}}
function downloadSalesPasteXls(rows, name='confirmed_sps_sales_rows') {{
  const cols = EXPORT_COLS.confirmed_sps_sales_rows;
  const table = `<table><thead><tr>${{cols.map(c => `<th>${{esc(c)}}</th>`).join('')}}</tr></thead><tbody>${{rows.map(row => `<tr>${{cols.map(c => `<td>${{esc(row[c] ?? '')}}</td>`).join('')}}</tr>`).join('')}}</tbody></table>`;
  const html = `<!doctype html><html><head><meta charset="utf-8"><style>td,th{{mso-number-format:"\\@";}}</style></head><body>${{table}}</body></html>`;
  const blob = new Blob(['\ufeff' + html], {{type:'application/vnd.ms-excel;charset=utf-8'}});
  const a = document.createElement('a');
  const stamp = new Date().toISOString().slice(0,19).replace(/[-:T]/g,'');
  a.href = URL.createObjectURL(blob);
  a.download = `${{name}}_${{stamp}}.xls`;
  a.click();
  URL.revokeObjectURL(a.href);
}}
async function copyCsv(items, cols=null) {{
  const csv = toCsv(items, cols);
  try {{
    await navigator.clipboard.writeText(csv);
    alert('Copied. You can paste into Excel.');
  }} catch {{
    const ta = document.createElement('textarea');
    ta.value = csv;
    document.body.appendChild(ta);
    ta.select();
    document.execCommand('copy');
    document.body.removeChild(ta);
    alert('Copied. You can paste into Excel.');
  }}
}}
function allocationRows() {{
  return Object.values(getAllocations()).filter(isDecisionRow);
}}
function getEmailSent() {{
  try {{ return JSON.parse(localStorage.getItem('scm_warehouse_email_sent') || '{{}}'); }}
  catch {{ return {{}}; }}
}}
function saveEmailSent(so, sent) {{
  const all = getEmailSent();
  if (sent) all[so] = {{so, sent_at:new Date().toISOString()}};
  else delete all[so];
  localStorage.setItem('scm_warehouse_email_sent', JSON.stringify(all));
}}
function normalizeFollowUpCustomer(customer) {{
  const text = String(customer || '');
  if (text.toLowerCase().includes('petco')) return 'Petco';
  if (text.toLowerCase().includes('petsmart')) return 'Petsmart';
  return text || 'Petco';
}}
function isExcludedSalesOrder(...values) {{
  const text = values.map(v => String(v || '').toLowerCase()).filter(Boolean).join(' ');
  if (text.includes('canada')) return true;
  if (text.includes('petco') && (text.includes('mx') || text.includes('mexico'))) return true;
  return ['petco-mx','petco mx','mx order','mexico order'].some(token => text.includes(token));
}}
function productNameForSku(sku, fallback='') {{
  return window.SCM_DATA.product_name_by_sku?.[sku] || fallback || '';
}}
function productCostForSku(sku) {{
  return Number(window.SCM_DATA.product_cost_by_sku?.[sku] || 0);
}}
function moneyValue(n) {{
  const v = Number(n || 0);
  return Number.isFinite(v) && v ? Math.round(v * 100) / 100 : '';
}}
function followUpOrderSkuKeys() {{
  return new Set((window.SCM_DATA.so_lines || []).map(x => `${{normalizeOrder(x.so)}}__${{normalizeSku(x.sku)}}`));
}}
function pendingConfirmedSpsRows() {{
  const existing = followUpOrderSkuKeys();
  return getConfirmedImports()
    .flatMap(batch => batch.new_rows || [])
    .filter(x => !isExcludedSalesOrder(x.customer, x.order, x.dc))
    .filter(x => !existing.has(`${{normalizeOrder(x.order)}}__${{normalizeSku(x.sku)}}`));
}}
function postedConfirmedSpsRows() {{
  const existing = followUpOrderSkuKeys();
  return getConfirmedImports()
    .flatMap(batch => batch.new_rows || [])
    .filter(x => !isExcludedSalesOrder(x.customer, x.order, x.dc))
    .filter(x => existing.has(`${{normalizeOrder(x.order)}}__${{normalizeSku(x.sku)}}`));
}}
function confirmedSpsSalesRows() {{
  const rows = pendingConfirmedSpsRows();
  return rows.map(x => ({{
    Date: x.order_date || '',
    Item: productNameForSku(x.sku, x.product),
    'Product Code': x.sku || '',
    'SO/CI': x.order || '',
    'Zoho Order': '',
    'ORD QTY': x.sps_qty || '',
    'Unit Price': moneyValue(x.unit_price),
    'Total Sales': moneyValue(Number(x.sps_qty || 0) * Number(x.unit_price || 0)),
    Customer: normalizeFollowUpCustomer(x.customer),
    'Delivery Center': x.dc || '',
    ETD: x.etd || '',
    ATD: '',
    'Product Cost': moneyValue(productCostForSku(x.sku)),
    'Total Cost': moneyValue(Number(x.sps_qty || 0) * productCostForSku(x.sku)),
    'Sales Order Update': 'Imported from SPS; paste into Follow Up Sales.'
  }}));
}}
function warehouseEmailRows(view) {{
  const sent = getEmailSent();
  return view.soBoard.map(so => {{
    const emailDue = addDays(so.required_arrival, -21);
    const days = diffDays(emailDue, window.SCM_DATA.generated_at);
    const isSent = !!sent[so.so];
    let noticeStatus = 'Not Due';
    if (isSent) noticeStatus = 'Sent';
    else if (days !== '' && days < 0) noticeStatus = 'Overdue';
    else if (days === 0) noticeStatus = 'Due Today';
    else if (days !== '' && days <= 7) noticeStatus = 'Upcoming';
    return {{
      notice_status: noticeStatus,
      so: so.so,
      customer: so.customer,
      delivery_center: so.delivery_center,
      required_arrival: so.required_arrival,
      email_due_date: emailDue,
      days_to_email: days,
      latest_ship: so.latest_ship,
      risk_status: so.status,
      line_count: so.line_count,
      issue_skus: so.issue_skus,
      sent_at: sent[so.so]?.sent_at || '',
      action: isSent ? 'Warehouse email marked sent.' : (days === '' ? 'Missing required arrival date.' : (days <= 0 ? 'Send warehouse notice and mark sent.' : 'Not due yet.')),
    }};
  }}).sort((a,b) => {{
    const ar = a.notice_status === 'Overdue' ? 0 : a.notice_status === 'Due Today' ? 1 : a.notice_status === 'Upcoming' ? 2 : a.notice_status === 'Not Due' ? 3 : 4;
    const br = b.notice_status === 'Overdue' ? 0 : b.notice_status === 'Due Today' ? 1 : b.notice_status === 'Upcoming' ? 2 : b.notice_status === 'Not Due' ? 3 : 4;
    return ar - br || String(a.email_due_date || '9999-12-31').localeCompare(String(b.email_due_date || '9999-12-31')) || String(a.so).localeCompare(String(b.so));
  }});
}}
function warehouseEmailPanel(rows, title='提前 21 天通知仓库', subtitle='提醒日期 = 客户要求到仓日 - 21 天。邮件发给美国仓库后，在这里标记已发送。') {{
  const body = rows.length ? rows.map(row => `<tr>
    <td>${{statusBadge(row.notice_status === 'Sent' || row.notice_status === 'Not Due' ? 'OK' : row.notice_status === 'Upcoming' ? 'Lead Time Watch' : 'Critical Risk')}} ${{esc(zhStatus(row.notice_status))}}</td>
    <td><button class="link" data-kind="so" data-value="${{esc(row.so)}}">${{esc(row.so)}}</button></td>
    <td>${{esc(row.customer)}}</td>
    <td>${{esc(row.delivery_center)}}</td>
    <td>${{esc(row.required_arrival)}}</td>
    <td>${{esc(row.email_due_date)}}</td>
    <td class="num">${{esc(row.days_to_email)}}</td>
    <td>${{statusBadge(row.risk_status)}}</td>
    <td class="num">${{esc(row.issue_skus)}}</td>
    <td><button class="email-toggle" data-so="${{esc(row.so)}}" data-sent="${{row.notice_status === 'Sent' ? '0' : '1'}}">${{row.notice_status === 'Sent' ? '撤销已发送' : '标记已发送'}}</button></td>
  </tr>`).join('') : '<tr><td class="empty" colspan="10">暂无记录</td></tr>';
  return `<section class="panel"><div class="panel-title"><h2>${{esc(title)}}</h2><p>${{esc(subtitle)}}</p></div>
    <div class="table-wrap"><table><thead><tr><th>通知状态</th><th>SO/CI</th><th>客户</th><th>客户仓</th><th>客户要求到仓日</th><th>邮件提醒日</th><th>剩余天数</th><th>风险</th><th>问题SKU数</th><th>确认</th></tr></thead><tbody>${{body}}</tbody></table></div></section>`;
}}
function renderTransitSettings() {{
  const target = document.getElementById('transit');
  if (!target) return;
  const settings = getTransitSettings();
  const dcMap = new Map();
  [...window.SCM_DATA.so_lines, ...confirmedSpsAsSalesLines()].forEach(line => {{
    const key = dcKey(line.customer, line.delivery_center);
    if (!line.delivery_center || dcMap.has(key)) return;
    dcMap.set(key, {{
      key,
      customer:line.customer,
      dc:line.delivery_center,
      state:extractState(line.delivery_center),
    }});
  }});
  const discoveredStates = new Set([...dcMap.values()].map(x => x.state).filter(Boolean));
  const stateList = [...new Set([...Object.keys(DEFAULT_TRANSIT_DAYS), ...Object.keys(settings.stateDays || {{}}), ...discoveredStates])].sort();
  const stateRows = stateList.map(st => {{
    const def = DEFAULT_TRANSIT_DAYS[st] ?? 7;
    const value = settings.stateDays?.[st] ?? def;
    const source = DEFAULT_TRANSIT_DAYS[st] ? 'Default' : 'Auto-added';
    return `<tr><td>${{esc(st)}}</td><td>${{esc(source)}}</td><td class="num">${{esc(def)}}</td><td><input class="small-input transit-state-input" data-state="${{esc(st)}}" type="number" min="0" max="30" value="${{esc(value)}}"></td></tr>`;
  }}).join('');
  const dcRows = [...dcMap.values()].sort((a,b)=>String(a.customer).localeCompare(String(b.customer))||String(a.dc).localeCompare(String(b.dc))).map(row => {{
    const stateDays = row.state ? (settings.stateDays?.[row.state] ?? DEFAULT_TRANSIT_DAYS[row.state] ?? 7) : 7;
    const value = settings.dcDays?.[row.key] ?? '';
    const current = value === '' ? transitDays(row.customer, row.dc) : Number(value);
    const source = dcStateSource(row.customer, row.dc, row.state, settings.dcDays?.[row.key] !== undefined && String(settings.dcDays?.[row.key]).trim() !== '');
    return `<tr><td>${{esc(row.customer)}}</td><td>${{esc(row.dc)}}</td><td>${{esc(row.state || 'Unknown')}}</td><td>${{esc(source)}}</td><td class="num">${{esc(stateDays)}}</td><td class="num">${{esc(current)}}</td><td><input class="small-input transit-dc-input" data-key="${{esc(row.key)}}" type="number" min="0" max="30" value="${{esc(value)}}" placeholder="${{esc(stateDays)}}"></td></tr>`;
  }}).join('');
  const autoDcCount = [...dcMap.values()].filter(x => !x.state || !DEFAULT_TRANSIT_DAYS[x.state]).length;
  target.innerHTML = `<section class="panel"><div class="panel-title"><h2>Transit Settings / 美仓到客户仓天数</h2><p>新客户/新客户仓会自动从 Follow Up 和已确认 SPS 新单里出现。未知仓默认按 7 天，Petsmart 自提为 0 天。</p></div>
    <div class="toolbox"><button id="resetTransitSettings">重置运输设置</button></div>
    <div class="detail-grid">${{mini('州规则数', stateList.length)}}${{mini('客户仓数', dcMap.size)}}${{mini('默认天数客户仓', autoDcCount)}}${{mini('缓冲天数', 1)}}</div>
    <div class="table-wrap"><table class="transit-table"><thead><tr><th>州</th><th>来源</th><th>默认天数</th><th>你的天数</th></tr></thead><tbody>${{stateRows}}</tbody></table></div>
    <div class="panel-title"><h2>单个客户仓运输天数覆盖</h2><p>空白表示使用州规则；只有某个客户仓需要特殊天数时再填写。</p></div>
    <div class="table-wrap"><table class="transit-table"><thead><tr><th>客户</th><th>客户仓</th><th>州</th><th>来源</th><th>州规则天数</th><th>当前天数</th><th>覆盖天数</th></tr></thead><tbody>${{dcRows || '<tr><td class="empty" colspan="7">没有客户仓</td></tr>'}}</tbody></table></div></section>`;
}}
function parseCsv(text) {{
  const rows = [];
  let row = [], cur = '', q = false;
  const LF = String.fromCharCode(10), CR = String.fromCharCode(13);
  let atFieldStart = true;
  for (let i=0;i<text.length;i++) {{
    const ch = text[i], next = text[i+1];
    if (ch === '"' && q && next === '"') {{ cur += '"'; i++; atFieldStart = false; continue; }}
    if (ch === '"' && !q && atFieldStart) {{ q = true; atFieldStart = false; continue; }}
    if (ch === '"' && q && (next === ',' || next === LF || next === CR || next === undefined)) {{ q = false; continue; }}
    if (ch === '"') {{ cur += ch; atFieldStart = false; continue; }}
    if (ch === ',' && !q) {{ row.push(cur); cur = ''; atFieldStart = true; continue; }}
    if ((ch === LF || ch === CR) && !q) {{
      if (ch === CR && next === LF) i++;
      row.push(cur); cur = '';
      if (row.some(x => String(x).trim() !== '')) rows.push(row);
      row = [];
      atFieldStart = true;
      continue;
    }}
    cur += ch;
    atFieldStart = false;
  }}
  row.push(cur);
  if (row.some(x => String(x).trim() !== '')) rows.push(row);
  if (!rows.length) return [];
  const headers = rows[0].map(x => String(x).trim());
  return rows.slice(1).map(r => Object.fromEntries(headers.map((h,i) => [h, r[i] ?? ''])));
}}
function pick(row, names) {{
  const keys = Object.keys(row);
  for (const name of names) {{
    const hit = keys.find(k => k.toLowerCase().replace(/[^a-z0-9]/g,'') === name.toLowerCase().replace(/[^a-z0-9]/g,''));
    if (hit && String(row[hit] ?? '').trim()) return String(row[hit]).trim();
  }}
  for (const name of names) {{
    const hit = keys.find(k => k.toLowerCase().includes(name.toLowerCase()));
    if (hit && String(row[hit] ?? '').trim()) return String(row[hit]).trim();
  }}
  return '';
}}
function normalizeOrder(v) {{ return String(v ?? '').trim(); }}
function normalizeSku(v) {{ return String(v ?? '').trim().toUpperCase(); }}
function normalizeUpc(v) {{ return String(v ?? '').replace(/[^0-9]/g,''); }}
function normalizeQty(v) {{
  const n = Number(String(v ?? '').replace(/,/g,'').trim());
  return Number.isFinite(n) ? n : 0;
}}
function normalizeDate(v) {{
  const s = String(v ?? '').trim();
  if (!s) return '';
  const m = s.match(new RegExp('(\\\\d{{4}})[-/](\\\\d{{1,2}})[-/](\\\\d{{1,2}})'));
  if (m) return `${{m[1]}}-${{String(m[2]).padStart(2,'0')}}-${{String(m[3]).padStart(2,'0')}}`;
  const d = new Date(s);
  return Number.isNaN(d.getTime()) ? s : d.toISOString().slice(0,10);
}}
function addDays(dateText, days) {{
  const d = new Date(`${{dateText}}T00:00:00`);
  if (Number.isNaN(d.getTime())) return '';
  d.setDate(d.getDate() + days);
  return d.toISOString().slice(0,10);
}}
function warehouseReadyDate(whEta) {{
  return whEta ? addDays(whEta, 7) : '';
}}
function diffDays(a, b) {{
  const da = new Date(`${{a}}T00:00:00`), db = new Date(`${{b}}T00:00:00`);
  if (Number.isNaN(da.getTime()) || Number.isNaN(db.getTime())) return '';
  return Math.round((da - db) / 86400000);
}}
const DEFAULT_TRANSIT_DAYS = {{CA:2,NV:3,AZ:3,OR:4,WA:4,UT:5,CO:5,TX:5,IL:7,GA:8,SC:8,NC:8,NJ:9,PA:9,NY:9,MD:9,VA:9,MA:10,CT:10,FL:10}};
const DEFAULT_DC_TRANSIT_DAYS = {{'MIRA LOMA':2,'RENO':3,'TEXAS':5,'JOLIET':7,'BRASELTON':8,'CRANBURY':9,'PSP DISTRIBUTION-SE':7,'PSP DISTRIBUTION-OR':8}};
const DEFAULT_DC_STATE_MAP = {{'MIRA LOMA':'CA','RENO':'NV','TEXAS':'TX','JOLIET':'IL','BRASELTON':'GA','CRANBURY':'NJ','PSP DISTRIBUTION-OR':'OR','PSP DISTRIBUTION-SE':'GA'}};
function getTransitSettings() {{
  try {{ return JSON.parse(localStorage.getItem('scm_transit_settings') || '{{"stateDays":{{}},"dcDays":{{}}}}'); }}
  catch {{ return {{stateDays:{{}}, dcDays:{{}}}}; }}
}}
function saveTransitSettings(settings) {{
  localStorage.setItem('scm_transit_settings', JSON.stringify(settings));
}}
function extractState(dc) {{
  const text = String(dc || '').toUpperCase();
  for (const [dcText, st] of Object.entries(DEFAULT_DC_STATE_MAP)) {{
    if (text.includes(dcText)) return st;
  }}
  const m = text.match(/(?:,|\\s)([A-Z]{{2}})\\s*$/);
  if (m && DEFAULT_TRANSIT_DAYS[m[1]]) return m[1];
  for (const st of Object.keys(DEFAULT_TRANSIT_DAYS)) {{
    if (text.includes(`, ${{st}}`) || text.endsWith(` ${{st}}`) || text.endsWith(st)) return st;
  }}
  return '';
}}
function dcKey(customer, dc) {{
  return `${{String(customer || '').trim().toUpperCase()}}__${{String(dc || '').trim().toUpperCase()}}`;
}}
function mappedDcState(dc) {{
  const text = String(dc || '').toUpperCase();
  for (const [dcText, st] of Object.entries(DEFAULT_DC_STATE_MAP)) {{
    if (text.includes(dcText)) return st;
  }}
  return '';
}}
function dcStateSource(customer, dc, state, hasOverride) {{
  if (String(customer || '').toLowerCase().includes('petsmart')) return hasOverride ? 'Override' : 'Pickup';
  if (hasOverride) return 'Override';
  if (mappedDcState(dc)) return 'Known DC map';
  if (state) return 'State in name';
  return 'Auto default';
}}
function transitDays(customer, dc) {{
  if (String(customer || '').toLowerCase().includes('petsmart')) return 0;
  const settings = getTransitSettings();
  const key = dcKey(customer, dc);
  if (settings.dcDays && String(settings.dcDays[key] ?? '').trim() !== '') return Number(settings.dcDays[key]);
  const text = String(dc || '').toUpperCase();
  for (const [dcText, days] of Object.entries(DEFAULT_DC_TRANSIT_DAYS)) {{
    if (text.includes(dcText)) return days;
  }}
  const states = {{...DEFAULT_TRANSIT_DAYS, ...(settings.stateDays || {{}})}};
  const mappedState = extractState(dc);
  if (mappedState && states[mappedState] !== undefined) return Number(states[mappedState]);
  for (const [st, days] of Object.entries(states)) {{
    if (text.includes(`, ${{st}}`) || text.endsWith(` ${{st}}`) || text.endsWith(st)) return days;
  }}
  return 7;
}}
function latestShipDate(customer, dc, etd) {{
  if (!etd) return '';
  return addDays(etd, -(transitDays(customer, dc) + 1));
}}
function suggestedCustomerDate(line, coverDate) {{
  if (!coverDate) return '';
  return addDays(coverDate, transitDays(line.customer, line.delivery_center) + 1);
}}
function latestFactoryDate(latestWhArrival) {{
  return latestWhArrival ? addDays(latestWhArrival, -37) : '';
}}
function withTransitDates(line) {{
  const latest = latestShipDate(line.customer, line.delivery_center, line.required_arrival);
  return {{
    ...line,
    domestic_transit: transitDays(line.customer, line.delivery_center),
    latest_ship: latest || line.latest_ship,
  }};
}}
function expectedUpcForSps(rawSku, sku) {{
  return normalizeUpc(window.SCM_DATA.customer_code_to_upc?.[rawSku] || window.SCM_DATA.product_upc_by_sku?.[sku] || '');
}}
function spsUpcIssue(rawSku, sku, upc) {{
  const expected = expectedUpcForSps(rawSku, sku);
  const actual = normalizeUpc(upc);
  if (!actual) return {{upc_status:'Missing UPC', expected_upc:expected, sps_upc:actual}};
  if (expected && actual !== expected) return {{upc_status:'UPC mismatch', expected_upc:expected, sps_upc:actual}};
  if (!expected && !window.SCM_DATA.upc_to_sku?.[actual]) return {{upc_status:'UPC not in Product List', expected_upc:expected, sps_upc:actual}};
  return {{upc_status:'OK', expected_upc:expected || actual, sps_upc:actual}};
}}
function spsCasePackIssue(sku, qty) {{
  const pack = Number(window.SCM_DATA.case_pack_by_sku?.[sku] || 0);
  const q = Number(qty || 0);
  if (!pack) return {{case_pack_status:'No Case Pack', case_pack:'', case_pack_remainder:''}};
  const remainder = q % pack;
  return {{
    case_pack_status: remainder === 0 ? 'OK' : 'Qty not case-pack multiple',
    case_pack:pack,
    case_pack_remainder:remainder,
  }};
}}
function spsLineFromRow(row, fileName) {{
  const order = normalizeOrder(pick(row, ['OrderKey','PO Number','SPS PO','PO','Purchase Order','Purchase Order Number','Retailers PO','Customer PO','Order','Order Number']));
  const rawSku = normalizeSku(pick(row, ['SKUKey','SKU','Buyers Catalog or Stock Keeping #','Buyer Part Number','Vendor Part Number','Item','Item Code','Product Code']));
  const upc = normalizeUpc(pick(row, ['UPC/EAN','UPC','GTIN']));
  const sku = normalizeSku(window.SCM_DATA.customer_code_to_sku?.[rawSku] || window.SCM_DATA.upc_to_sku?.[upc] || rawSku);
  const qty = normalizeQty(pick(row, ['SPS Qty','Qty Ordered','Quantity','Qty','Order Qty','ORD QTY','Quantity Ordered']));
  const unitPrice = normalizeQty(pick(row, ['Unit Price','Price','Unit Cost','Cost']));
  const customer = pick(row, ['Partner','Trading Partner','Customer','Retailer']);
  const dc = pick(row, ['Ship To Name','Ship To','Delivery Center','Ship To Location','Location']);
  const orderDate = normalizeDate(pick(row, ['PO Date','Order Date','Date']));
  const requested = normalizeDate(pick(row, ['Delivery Dates','Requested Delivery Date','Delivery Date','Requested Date','ETD']));
  const shipDate = normalizeDate(pick(row, ['Ship Dates','Ship Date','Requested Ship Date']));
  const product = pick(row, ['SPS Product Name','Product Name','Description','Item Description']);
  const recordType = pick(row, ['Record Type']);
  return {{order, raw_sku: rawSku, sku, upc, qty, unit_price:unitPrice, customer, dc, orderDate, requested, shipDate, product, recordType, source_file:fileName}};
}}
let lastSpsDiff = [];
let lastSpsRiskPreview = [];
function compareSps(lines) {{
  const existing = new Map();
  window.SCM_DATA.so_lines.forEach(l => {{
    const key = `${{normalizeOrder(l.so)}}__${{normalizeSku(l.sku)}}`;
    existing.set(key, (existing.get(key) || 0) + Number(l.qty || 0));
  }});
  const grouped = new Map();
  lines.forEach(l => {{
    const key = `${{l.order}}__${{l.sku}}`;
    if (!l.order || !l.sku || !l.qty) {{
      grouped.set(`${{key}}__bad__${{grouped.size}}`, {{...l, issue:'Missing required fields'}});
      return;
    }}
    const prev = grouped.get(key) || {{...l, qty:0, unit_price:l.unit_price || 0, upc:l.upc || ''}};
    prev.qty += l.qty;
    if (!prev.unit_price && l.unit_price) prev.unit_price = l.unit_price;
    if (!prev.upc && l.upc) prev.upc = l.upc;
    grouped.set(key, prev);
  }});
  return [...grouped.values()].map(l => {{
    if (l.issue) return l;
    const key = `${{l.order}}__${{l.sku}}`;
    const fuQty = existing.get(key) || 0;
    const issue = !fuQty ? 'New in SPS' : (Math.abs(fuQty - l.qty) > 0.0001 ? 'Qty mismatch' : 'Already in Follow Up');
    const upcCheck = spsUpcIssue(l.raw_sku, l.sku, l.upc);
    const casePackCheck = spsCasePackIssue(l.sku, l.qty);
    return {{issue, order:l.order, raw_sku:l.raw_sku, sku:l.sku, product:l.product, customer:l.customer, dc:l.dc, order_date:l.orderDate, etd:l.requested, ship_date:l.shipDate, sps_qty:l.qty, unit_price:l.unit_price || '', followup_qty:fuQty, qty_diff:l.qty - fuQty, ...upcCheck, ...casePackCheck, source_file:l.source_file}};
  }}).sort((a,b) => String(a.issue).localeCompare(String(b.issue)) || String(a.order).localeCompare(String(b.order)) || String(a.sku).localeCompare(String(b.sku)));
}}
function simulateNewOrderRisk(diff) {{
  const newLines = diff.filter(x => x.issue === 'New in SPS').map(x => ({{
    status:'New SPS',
    so:x.order,
    sku:x.sku,
    product:x.product,
    qty:Number(x.sps_qty || 0),
    customer:x.customer || 'Petco',
    delivery_center:x.dc,
    required_arrival:x.etd,
    latest_ship:latestShipDate(x.customer || 'Petco', x.dc, x.etd),
    source:'new',
    raw:x,
  }}));
  const bySku = new Map();
  for (const l of window.SCM_DATA.so_lines) {{
    if (!bySku.has(l.sku)) bySku.set(l.sku, []);
    bySku.get(l.sku).push({{...l, source:'existing', qty:Number(l.qty || 0)}});
  }}
  for (const l of newLines) {{
    if (!bySku.has(l.sku)) bySku.set(l.sku, []);
    bySku.get(l.sku).push(l);
  }}
  const results = [];
  for (const [sku, lines] of bySku.entries()) {{
    if (!newLines.some(x => x.sku === sku)) continue;
    const skuInfo = skuInfoFor(sku);
    const existingDemand = Number(skuInfo.open_demand || 0);
    const newSpsDemand = newLines.filter(x => x.sku === sku).reduce((sum, x) => sum + Number(x.qty || 0), 0);
    const onhand = Number(skuInfo.current_onhand || 0);
    const gapAfterSps = onhand - existingDemand - newSpsDemand;
    const why = `Onhand ${{onhand}} - existing demand ${{existingDemand}} - new SPS demand ${{newSpsDemand}} = ${{gapAfterSps}}`;
    let stock = Number(skuInfo.current_onhand || 0);
    const pos = usableFuturePos(sku).map(x => ({{...x, remaining:Number(x.qty || 0)}})).sort((a,b) => String(a.available).localeCompare(String(b.available)) || String(a.po).localeCompare(String(b.po)));
    const ordered = lines.slice().sort((a,b) => String(a.latest_ship || '9999-12-31').localeCompare(String(b.latest_ship || '9999-12-31')) || String(a.required_arrival || '9999-12-31').localeCompare(String(b.required_arrival || '9999-12-31')) || String(a.so).localeCompare(String(b.so)));
    for (const line of ordered) {{
      let need = Number(line.qty || 0);
      const stockUsed = Math.min(Math.max(stock, 0), need);
      stock -= stockUsed;
      need -= stockUsed;
      const covers = [];
      for (const po of pos) {{
        if (need <= 0) break;
        if (po.remaining <= 0) continue;
        const take = Math.min(po.remaining, need);
        po.remaining -= take;
        need -= take;
        covers.push({{po:po.po, qty:take, available:po.available, reliability:po.reliability}});
      }}
      if (line.source !== 'new') continue;
      const coverDate = covers.map(x => x.available).sort().slice(-1)[0] || '';
      const warehouseReady = warehouseReadyDate(coverDate);
      let status = 'OK';
      let action = 'Can be covered by current stock.';
      if (need > 0) {{
        status = 'Critical Risk';
        action = `Short ${{need}} units after current stock and known POs.`;
      }} else if (coverDate && line.latest_ship && coverDate > line.latest_ship) {{
        status = 'Critical Risk';
        action = `Covered late by ${{diffDays(coverDate, line.latest_ship)}} days. Need earlier PO/ship plan.${{isPlannedCover(covers) ? ' Cover is planned ETA; confirm sailing date.' : ''}}`;
      }} else if (warehouseReady && line.latest_ship && warehouseReady > line.latest_ship) {{
        status = 'Warehouse Prep Risk';
        action = `Warehouse needs 7 days after arrival; suggest customer date ${{suggestedCustomerDate(line, warehouseReady)}}.`;
      }} else if (covers.length) {{
        status = isPlannedCover(covers) ? 'Planned PO Watch' : 'OK';
        action = isPlannedCover(covers) ? 'Covered only if planned ETA holds; confirm sailing date before promising.' : 'Covered by confirmed incoming PO on time.';
      }}
      const suggestedDate = (status === 'Warehouse Prep Risk' && warehouseReady) ? suggestedCustomerDate(line, warehouseReady) : (coverDate && line.latest_ship && coverDate > line.latest_ship ? suggestedCustomerDate(line, coverDate) : '');
      if (status === 'Critical Risk' && suggestedDate && line.required_arrival && suggestedDate <= line.required_arrival) {{
        status = isPlannedCover(covers) ? 'Planned PO Watch' : 'OK';
        action = isPlannedCover(covers) ? 'Covered if planned ETA holds; confirm sailing date.' : 'Covered by incoming PO under current customer date.';
      }} else if (status === 'Critical Risk' && suggestedDate) {{
        action += ` Suggest customer date ${{suggestedDate}}.`;
      }}
      results.push({{
        status, order:line.so, sku:line.sku, product:line.product, customer:line.customer, dc:line.delivery_center,
        etd:line.required_arrival, required_arrival:line.required_arrival, latest_ship:line.latest_ship, qty:line.qty,
        suggested_customer_date:suggestedDate,
        onhand, existing_demand: existingDemand, new_sps_demand: newSpsDemand, gap_after_sps: gapAfterSps,
        cover:covers.map(c => `${{c.po}}: ${{c.qty}} @ ${{c.available}} (${{c.reliability}})`).join('; ') || 'Current stock',
        uncovered:need, why, action
      }});
    }}
  }}
  return results.sort((a,b) => statusRank(a.status) - statusRank(b.status) || String(a.latest_ship).localeCompare(String(b.latest_ship)) || String(a.order).localeCompare(String(b.order)));
}}
function buildSpsSkuCoverageSummary(diff) {{
  const newRows = diff.filter(x => x.issue === 'New in SPS');
  const skus = [...new Set(newRows.map(x => x.sku).filter(Boolean))];
  return skus.map(sku => {{
    const skuInfo = skuInfoFor(sku);
    const onhand = Number(skuInfo.current_onhand || 0);
    const existingDemand = Number(skuInfo.open_demand || 0);
    const newSpsDemand = newRows.filter(x => x.sku === sku).reduce((sum, x) => sum + Number(x.sps_qty || 0), 0);
    const totalDemand = existingDemand + newSpsDemand;
    const pos = usableFuturePos(sku).map(x => ({{...x, qty:Number(x.qty || 0)}})).sort((a,b) => String(a.available || '9999-12-31').localeCompare(String(b.available || '9999-12-31')) || String(a.po).localeCompare(String(b.po)));
    const confirmedIncoming = pos.filter(x => !isPlannedReliability(x.reliability)).reduce((sum, x) => sum + x.qty, 0);
    const plannedIncoming = pos.filter(x => isPlannedReliability(x.reliability)).reduce((sum, x) => sum + x.qty, 0);
    const reliableSupply = onhand + confirmedIncoming;
    const totalSupply = reliableSupply + plannedIncoming;
    const reliableGap = reliableSupply - totalDemand;
    const totalGap = totalSupply - totalDemand;
    const status = reliableGap >= 0 ? 'OK' : (totalGap >= 0 ? 'Planned PO Watch' : 'Critical Risk');
    const action = reliableGap >= 0
      ? 'Confirmed stock/incoming can cover total demand; no add-on order needed based on current data.'
      : (totalGap >= 0 ? `Need planned PO to hold; reliable supply is short by ${{Math.abs(reliableGap)}}.` : `Still short ${{Math.abs(totalGap)}} after all known POs; consider add-on PO or customer date change.`);
    const firstCover = (() => {{
      let cumulative = onhand;
      for (const po of pos) {{
        cumulative += po.qty;
        if (cumulative >= totalDemand) return `${{po.po || '(blank PO)'}} @ ${{po.available || ''}} (${{po.reliability || ''}})`;
      }}
      return totalSupply >= totalDemand ? 'Covered after known POs' : '';
    }})();
    return {{
      status,
      sku,
      product:skuInfo.product || newRows.find(x => x.sku === sku)?.product || '',
      onhand,
      existing_demand:existingDemand,
      new_sps_demand:newSpsDemand,
      total_demand:totalDemand,
      confirmed_incoming:confirmedIncoming,
      planned_incoming:plannedIncoming,
      reliable_gap:reliableGap,
      total_gap:totalGap,
      first_cover:firstCover,
      action,
    }};
  }}).sort((a,b) => statusRank(a.status)-statusRank(b.status) || Number(a.total_gap)-Number(b.total_gap) || String(a.sku).localeCompare(String(b.sku)));
}}
function buildSpsPoCoverageLadder(diff) {{
  const newRows = diff.filter(x => x.issue === 'New in SPS');
  const skus = [...new Set(newRows.map(x => x.sku).filter(Boolean))];
  const rows = [];
  for (const sku of skus) {{
    const skuInfo = skuInfoFor(sku);
    const onhand = Number(skuInfo.current_onhand || 0);
    const existingDemand = Number(skuInfo.open_demand || 0);
    const newSpsDemand = newRows.filter(x => x.sku === sku).reduce((sum, x) => sum + Number(x.sps_qty || 0), 0);
    const totalDemand = existingDemand + newSpsDemand;
    let cumulative = onhand;
    rows.push({{
      status:cumulative >= totalDemand ? 'OK' : 'Lead Time Watch',
      sku,
      product:skuInfo.product || '',
      step:'Current stock',
      eta:'',
      reliability:'Onhand',
      step_qty:onhand,
      cumulative_supply:cumulative,
      total_demand:totalDemand,
      remaining_gap:cumulative - totalDemand,
      conclusion:cumulative >= totalDemand ? 'Current stock covers all demand.' : `Need ${{totalDemand - cumulative}} more units from incoming PO or add-on PO.`,
    }});
    const pos = usableFuturePos(sku).map(x => ({{...x, qty:Number(x.qty || 0)}})).sort((a,b) => String(a.available || '9999-12-31').localeCompare(String(b.available || '9999-12-31')) || String(a.po).localeCompare(String(b.po)));
    for (const po of pos) {{
      cumulative += Number(po.qty || 0);
      const gap = cumulative - totalDemand;
      rows.push({{
        status:gap >= 0 ? 'OK' : (isPlannedReliability(po.reliability) ? 'Planned PO Watch' : 'Lead Time Watch'),
        sku,
        product:skuInfo.product || '',
        step:po.po || '(blank PO)',
        eta:po.available || '',
        reliability:po.reliability || '',
        step_qty:po.qty,
        cumulative_supply:cumulative,
        total_demand:totalDemand,
        remaining_gap:gap,
        conclusion:gap >= 0 ? 'Cumulative supply covers all current + new SPS demand.' : `Still short ${{Math.abs(gap)}} after this PO.`,
      }});
    }}
  }}
  return rows;
}}
function statusRank(status) {{ return status === 'Critical Risk' ? 0 : status === 'Planned PO Watch' ? 1 : status === 'Lead Time Watch' ? 2 : 3; }}
function renderSpsDiff(diff) {{
  lastSpsDiff = diff;
  lastSpsRiskPreview = simulateNewOrderRisk(diff);
  const spsSkuCoverage = buildSpsSkuCoverageSummary(diff);
  const spsPoLadder = buildSpsPoCoverageLadder(diff);
  const counts = diff.reduce((acc,x) => (acc[x.issue]=(acc[x.issue]||0)+1, acc), {{}});
  const riskCounts = lastSpsRiskPreview.reduce((acc,x) => (acc[x.status]=(acc[x.status]||0)+1, acc), {{}});
  const upcIssueCount = diff.filter(x => x.upc_status && x.upc_status !== 'OK').length;
  const casePackIssueCount = diff.filter(x => x.case_pack_status && x.case_pack_status !== 'OK').length;
  document.getElementById('spsSummary').innerHTML =
    mini('SPS新增', counts['New in SPS'] || 0) +
    mini('数量不一致', counts['Qty mismatch'] || 0) +
    mini('已在 Follow Up', counts['Already in Follow Up'] || 0) +
    mini('缺少字段', counts['Missing required fields'] || 0) +
    mini('UPC异常', upcIssueCount) +
    mini('箱规异常', casePackIssueCount) +
    mini('新高风险', riskCounts['Critical Risk'] || 0) +
    mini('新预计在途', riskCounts['Planned PO Watch'] || 0) +
    mini('新观察', riskCounts['Lead Time Watch'] || 0);
  document.getElementById('spsRiskPreview').innerHTML =
    simplePanel('SPS 新单 SKU 总体覆盖', spsSkuCoverage, [['status','状态'],['sku','SKU'],['product','产品'],['onhand','当前库存'],['existing_demand','已有未交货需求'],['new_sps_demand','SPS新单需求'],['total_demand','总需求'],['confirmed_incoming','确认在途'],['planned_incoming','预计在途'],['reliable_gap','可靠缺口'],['total_gap','总缺口'],['first_cover','首次完全覆盖'],['action','处理建议']], '判断当前库存 + PO 供应是否能覆盖“已有需求 + SPS 新单需求”。可靠缺口不含预计在途；总缺口包含所有已知 PO。') +
    simplePanel('SPS PO 覆盖阶梯', spsPoLadder, [['status','状态'],['sku','SKU'],['product','产品'],['step','库存/PO'],['eta','预计到仓'],['reliability','可靠性'],['step_qty','数量'],['cumulative_supply','累计供应'],['total_demand','总需求'],['remaining_gap','该节点后缺口'],['conclusion','结论']], '按 SKU 从上往下看：每个 PO 到仓后，累计供应是否已经覆盖总需求。') +
    simplePanel('SPS 新单逐行影响', lastSpsRiskPreview, [['status','状态'],['order','SO/PO'],['sku','SKU'],['product','产品'],['customer','客户'],['dc','客户仓'],['required_arrival','客户要求到仓日'],['latest_ship','最晚美仓发货日'],['suggested_customer_date','建议客户交期'],['qty','数量'],['onhand','当前库存'],['existing_demand','已有需求'],['new_sps_demand','SPS新单需求'],['gap_after_sps','导入后缺口'],['cover','覆盖来源'],['uncovered','未覆盖数量'],['why','说明'],['action','处理建议']], '逐行解释上方 SKU 汇总后的订单影响。');
  document.getElementById('spsResults').innerHTML = simpleTable(diff, [['issue','问题'],['upc_status','UPC检查'],['case_pack_status','箱规检查'],['order','SO/PO'],['raw_sku','客户编码'],['sku','SKU'],['sps_upc','SPS UPC'],['expected_upc','应有UPC'],['case_pack','箱规'],['case_pack_remainder','箱规余数'],['product','产品'],['customer','客户'],['dc','客户仓'],['order_date','订单日期'],['etd','ETD'],['sps_qty','SPS数量'],['unit_price','单价'],['followup_qty','Follow Up数量'],['qty_diff','数量差异'],['source_file','文件']]);
}}
function getConfirmedImports() {{
  try {{ return JSON.parse(localStorage.getItem('scm_confirmed_sps_imports') || '[]'); }}
  catch {{ return []; }}
}}
function confirmedSpsAsSalesLines() {{
  const rows = pendingConfirmedSpsRows();
  return rows.map((x, idx) => ({{
    status:'New SPS',
    system_status:'New SPS',
    so:x.order,
    customer:x.customer || 'Petco',
    delivery_center:x.dc,
    sku:x.sku,
    product:x.product,
    qty:Number(x.sps_qty || 0),
    required_arrival:x.etd,
    latest_ship:latestShipDate(x.customer || 'Petco', x.dc, x.etd),
    domestic_transit:transitDays(x.customer || 'Petco', x.dc),
    stock_before:'',
    stock_used:0,
    cover:'New SPS confirmed import',
    cover_pos:[],
    cover_eta:'',
    uncovered:0,
    risk_qty:Number(x.sps_qty || 0),
    action:'Confirmed SPS line included in local dashboard calculation.',
    row:`SPS-${{idx+1}}`,
    lead_weeks:'',
    lead_due:'',
    source:'confirmed_sps',
    raw:x,
  }}));
}}
function renderImportNotice() {{
  const imports = getConfirmedImports();
  const panel = document.getElementById('importNotice');
  if (!panel) return;
  panel.style.display = 'none';
  const spsTab = document.getElementById('spsResults');
  if (!spsTab) return;
  if (!imports.length) return;
  const rows = imports.flatMap(x => x.new_rows || []);
  const pendingRows = pendingConfirmedSpsRows();
  const postedRows = postedConfirmedSpsRows();
  const latest = imports[imports.length - 1];
  const confirmedHtml = `<section class="panel"><div class="panel-title"><h2>已确认 SPS 新订单</h2><p>最近确认时间：${{esc(latest.confirmed_at)}}。待回填行已纳入本地计算；已在 Follow Up 找到的行会排除，避免需求重复。</p></div>
    <div class="detail-grid">${{mini('确认行数', rows.length)}}${{mini('待回填', pendingRows.length)}}${{mini('已在 Follow Up', postedRows.length)}}${{mini('是否需导出', pendingRows.length ? '是' : '否')}}</div>
    ${{simpleTable(pendingRows, [['issue','问题'],['order','SO/PO'],['sku','SKU'],['product','产品'],['customer','客户'],['dc','客户仓'],['etd','ETD'],['sps_qty','数量'],['source_file','文件']])}}
    ${{postedRows.length ? simplePanel('已回填到 Follow Up，不再重复计算', postedRows, [['order','SO/PO'],['sku','SKU'],['product','产品'],['customer','客户'],['dc','客户仓'],['etd','ETD'],['sps_qty','数量']], '这些已确认 SPS 行仅保留历史记录，不再参与需求计算和导出。') : ''}}</section>`;
  if (!document.getElementById('confirmedSpsLines')) {{
    const wrap = document.createElement('div');
    wrap.id = 'confirmedSpsLines';
    wrap.innerHTML = confirmedHtml;
    spsTab.parentElement?.insertBefore(wrap, spsTab);
  }} else {{
    document.getElementById('confirmedSpsLines').innerHTML = confirmedHtml;
  }}
}}
async function loadSpsFiles(files) {{
  const all = [];
  for (const file of files) {{
    if (file.name.toLowerCase().endsWith('.zip')) {{
      alert('This browser version cannot read zip directly yet. Please unzip SPS first and select the CSV files.');
      continue;
    }}
    const text = await file.text();
    let context = {{}};
    parseCsv(text).forEach(row => {{
      for (const key of ['PO Number','PO Date','Delivery Dates','Requested Delivery Date','Ship Dates','Ship To Name','Ship To Location','Partner']) {{
        if (row[key]) context[key] = row[key];
        else if (context[key]) row[key] = context[key];
      }}
      const line = spsLineFromRow(row, file.name);
      if (line.recordType && line.recordType !== 'D') return;
      if (!line.raw_sku && !line.qty) return;
      all.push(line);
    }});
  }}
  renderSpsDiff(compareSps(all));
}}
document.getElementById('search').addEventListener('input', (e) => {{
  activeFilter = null;
  applyGlobalFilter();
}});
document.getElementById('clearActiveFilter')?.addEventListener('click', () => {{
  activeFilter = null;
  const search = document.getElementById('search');
  if (search) search.value = '';
  applyGlobalFilter();
}});
document.addEventListener('click', (e) => {{
  const btn = e.target.closest('#clearBolInlineFilter');
  if (!btn) return;
  activeFilter = null;
  const search = document.getElementById('search');
  if (search) search.value = '';
  applyGlobalFilter();
  document.getElementById('bolSummary')?.scrollIntoView({{behavior:'smooth', block:'start'}});
}});
document.addEventListener('input', (e) => {{
  const el = e.target.closest('.alloc-input');
  if (!el) return;
  if (el.dataset.field === 'assign_qty') capManualStockInput(el);
  saveAllocation(el.dataset.sku, el.dataset.so, el.dataset.field, el.value);
  if (el.dataset.field === 'assign_qty' && String(el.value).trim() !== '') {{
    const row = el.closest('tr');
    const select = row?.querySelector('.alloc-input[data-field="assign"]');
    if (select) select.value = 'manual';
  }}
  updateAllocationSummary(el.dataset.sku);
  renderDynamicViews();
}});
document.addEventListener('change', (e) => {{
  const el = e.target.closest('.alloc-input');
  if (!el) return;
  if (el.dataset.field === 'assign_qty') capManualStockInput(el);
  saveAllocation(el.dataset.sku, el.dataset.so, el.dataset.field, el.value);
  renderDynamicViews();
  showDetail('sku', el.dataset.sku);
}});
document.addEventListener('click', (e) => {{
  const btn = e.target.closest('.clear-sku-allocation');
  if (!btn) return;
  const sku = btn.dataset.sku;
  const all = getAllocations();
  for (const key of Object.keys(all)) {{
    if (all[key].sku === sku) delete all[key];
  }}
  localStorage.setItem('scm_manual_allocations', JSON.stringify(all));
  renderDynamicViews();
  showDetail('sku', sku);
}});
document.addEventListener('click', (e) => {{
  const btn = e.target.closest('.email-toggle');
  if (!btn) return;
  saveEmailSent(btn.dataset.so, btn.dataset.sent === '1');
  renderDynamicViews();
}});
document.addEventListener('change', (e) => {{
  const stateInput = e.target.closest('.transit-state-input');
  const dcInput = e.target.closest('.transit-dc-input');
  if (!stateInput && !dcInput) return;
  const settings = getTransitSettings();
  settings.stateDays = settings.stateDays || {{}};
  settings.dcDays = settings.dcDays || {{}};
  if (stateInput) {{
    const st = stateInput.dataset.state;
    const n = Number(stateInput.value);
    if (Number.isFinite(n) && n >= 0) settings.stateDays[st] = n;
  }}
  if (dcInput) {{
    const key = dcInput.dataset.key;
    const raw = String(dcInput.value ?? '').trim();
    if (raw === '') delete settings.dcDays[key];
    else {{
      const n = Number(raw);
      if (Number.isFinite(n) && n >= 0) settings.dcDays[key] = n;
    }}
  }}
  saveTransitSettings(settings);
  if (lastSpsDiff.length) renderSpsDiff(lastSpsDiff);
  renderDynamicViews();
}});
document.addEventListener('click', (e) => {{
  const btn = e.target.closest('#resetTransitSettings');
  if (!btn) return;
  localStorage.removeItem('scm_transit_settings');
  if (lastSpsDiff.length) renderSpsDiff(lastSpsDiff);
  renderDynamicViews();
}});
document.getElementById('spsFiles')?.addEventListener('change', (e) => loadSpsFiles([...e.target.files]));
document.getElementById('uploadFollowup')?.addEventListener('click', async () => {{
  const status = document.getElementById('followupUploadStatus');
  const file = document.getElementById('followupFile')?.files?.[0];
  if (!file) {{ status.textContent = '请先选择一份 Follow Up 表格。'; return; }}
  if (location.protocol === 'file:') {{
    status.textContent = '直接打开的 HTML 不能上传文件，请使用在线看板网址。';
    return;
  }}
  const form = new FormData();
  form.append('followup', file);
  status.textContent = '正在上传并重新计算...';
  try {{
    const res = await fetch('/api/upload_followup', {{
      method:'POST',
      body:form,
      headers:{{'Accept':'application/json', 'X-Requested-With':'SCM-Dashboard'}}
    }});
    const text = await res.text();
    let data = null;
    try {{
      data = JSON.parse(text);
    }} catch (parseErr) {{
      if (text.trim().startsWith('<')) {{
        throw new Error('服务器返回了网页而不是上传结果，请确认你已登录在线看板，并且网址不是 file://。');
      }}
      throw new Error(text.slice(0, 300) || '服务器返回格式不正确。');
    }}
    if (!res.ok || !data.ok) throw new Error(data.error || '上传失败');
    status.textContent = '重新计算完成，正在刷新看板...';
    location.href = data.url || '/';
  }} catch (err) {{
    status.textContent = `上传失败：${{err.message}}`;
  }}
}});
document.getElementById('confirmSpsImport')?.addEventListener('click', () => {{
  if (!lastSpsDiff.length) {{ alert('Import SPS files first.'); return; }}
  const upcIssues = lastSpsDiff.filter(x => x.upc_status && x.upc_status !== 'OK').length;
  const packIssues = lastSpsDiff.filter(x => x.case_pack_status && x.case_pack_status !== 'OK').length;
  if ((upcIssues || packIssues) && !confirm(`SPS import has ${{upcIssues}} UPC issue(s) and ${{packIssues}} case-pack issue(s). Confirm import anyway?`)) return;
  const newRows = lastSpsDiff.filter(x => x.issue === 'New in SPS');
  const imports = getConfirmedImports();
  imports.push({{
    confirmed_at: new Date().toISOString(),
    diff: lastSpsDiff,
    new_rows: newRows,
    risk: lastSpsRiskPreview,
  }});
  localStorage.setItem('scm_confirmed_sps_imports', JSON.stringify(imports));
  renderImportNotice();
  renderDynamicViews();
  alert(`Confirmed ${{newRows.length}} new SPS line(s). Dashboard recalculated locally.`);
}});
document.getElementById('clearSps')?.addEventListener('click', () => {{
  lastSpsDiff = [];
  lastSpsRiskPreview = [];
  document.getElementById('spsSummary').innerHTML = '';
  document.getElementById('spsRiskPreview').innerHTML = '';
  document.getElementById('spsResults').innerHTML = '<div class="hint">No SPS files imported yet.</div>';
}});
document.getElementById('clearConfirmedImports')?.addEventListener('click', () => {{
  localStorage.removeItem('scm_confirmed_sps_imports');
  renderImportNotice();
  document.getElementById('confirmedSpsLines')?.remove();
  renderDynamicViews();
  alert('Confirmed SPS imports cleared from this browser.');
}});
document.getElementById('downloadSpsDiff')?.addEventListener('click', () => downloadCsv('sps_difference_for_review.csv', lastSpsDiff));
document.getElementById('copySpsNew')?.addEventListener('click', () => {{
  const confirmedRows = confirmedSpsSalesRows();
  const rows = confirmedRows.length ? confirmedRows : lastSpsDiff.filter(x => x.issue === 'New in SPS').map(x => ({{
    Date: x.order_date, Item: productNameForSku(x.sku, x.product), 'Product Code': x.sku, 'SO/CI': x.order, 'Zoho Order': '',
    'ORD QTY': x.sps_qty, 'Unit Price': moneyValue(x.unit_price), 'Total Sales': moneyValue(Number(x.sps_qty || 0) * Number(x.unit_price || 0)), Customer: normalizeFollowUpCustomer(x.customer),
    'Delivery Center': x.dc, ETD: x.etd, ATD: '', 'Product Cost': moneyValue(productCostForSku(x.sku)), 'Total Cost': moneyValue(Number(x.sps_qty || 0) * productCostForSku(x.sku)),
    'Sales Order Update': 'Imported from SPS; paste into Follow Up Sales.'
  }}));
  copyCsv(rows, EXPORT_COLS.confirmed_sps_sales_rows);
}});
document.getElementById('downloadSpsSalesXls')?.addEventListener('click', () => {{
  const rows = confirmedSpsSalesRows().length ? confirmedSpsSalesRows() : lastSpsDiff.filter(x => x.issue === 'New in SPS').map(x => ({{
    Date: x.order_date, Item: productNameForSku(x.sku, x.product), 'Product Code': x.sku, 'SO/CI': x.order, 'Zoho Order': '',
    'ORD QTY': x.sps_qty, 'Unit Price': moneyValue(x.unit_price), 'Total Sales': moneyValue(Number(x.sps_qty || 0) * Number(x.unit_price || 0)), Customer: normalizeFollowUpCustomer(x.customer),
    'Delivery Center': x.dc, ETD: x.etd, ATD: '', 'Product Cost': moneyValue(productCostForSku(x.sku)), 'Total Cost': moneyValue(Number(x.sps_qty || 0) * productCostForSku(x.sku)),
    'Sales Order Update': 'Imported from SPS; paste into Follow Up Sales.'
  }}));
  downloadSalesPasteXls(rows, 'sps_sales_paste_rows');
}});
document.getElementById('downloadConfirmedSpsXls')?.addEventListener('click', () => downloadSalesPasteXls(confirmedSpsSalesRows(), 'confirmed_sps_sales_rows'));
document.querySelectorAll('[data-export]').forEach(btn => btn.addEventListener('click', () => {{
  const key = btn.dataset.export;
  const view = currentViewData();
  const dynamic = {{
    action_queue:view.actionQueue,
    so_board:view.soBoard,
    so_lines:view.lines,
    sku_summary:buildSkuShortageRows(view),
    logistics_actions:logisticsActionRows(view),
    inbound_need_by:buildInboundNeedByRows(view),
    manual_audit:view.audit,
    warehouse_email:warehouseEmailRows(view),
    confirmed_sps_sales_rows:confirmedSpsSalesRows(),
  }};
  downloadCsv(`${{key}}.csv`, dynamic[key] || window.SCM_DATA[key] || [], EXPORT_COLS[key]);
}}));
document.getElementById('downloadAllocations')?.addEventListener('click', () => downloadCsv('manual_allocation_decisions.csv', allocationRows()));
renderImportNotice();
renderDynamicViews();
document.addEventListener('click', (e) => {{
  const btn = e.target.closest('button.link');
  if (!btn) return;
  if (btn.dataset.kind === 'bol') {{
    filterAndFocusBol(btn.dataset.value);
    return;
  }}
  showDetail(btn.dataset.kind, btn.dataset.value);
}});
</script>
</body>
</html>"""
    return html


if __name__ == "__main__":
    main()
